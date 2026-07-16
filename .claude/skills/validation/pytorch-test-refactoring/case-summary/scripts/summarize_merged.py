#!/usr/bin/env python3
"""Add a case_count column to the merged xlsx and write a summary sheet.

Consumes case_counts.tsv (from collect_case_counts.py). Adds a case_count
column keyed by (file, class), then computes four analyses and writes them to a
new 'summary' sheet:

  #1 category distribution by CLASS count + xpu_enabled % within device_agnostic
  #2 same as #1 but restricted to ut_priority in {P0,P1,P2,P3, blank}
  #4a analysis #1 recomputed weighting by CASE count
  #4b analysis #2 recomputed weighting by CASE count
"""

import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

PRIORITY_SUBSET = {"P0", "P1", "P2", "P3", ""}
CATEGORIES = ["device_unrelated", "device_agnostic", "device_specific"]


def load_case_counts(tsv: Path) -> dict[tuple[str, str], int]:
    out: dict[tuple[str, str], int] = {}
    if not tsv.exists():
        return out
    with open(tsv) as f:
        next(f, None)
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) != 3:
                continue
            file, cls, n = parts
            out[(file, cls)] = int(n)
    return out


def norm_priority(val) -> str:
    s = "" if val is None else str(val).strip()
    return s if s in {"P0", "P1", "P2", "P3"} else ("" if s == "" else s)


def pct(n: int, d: int) -> float:
    return round(100.0 * n / d, 2) if d else 0.0


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default="agent_space_xpu/test_class_classification_merged.xlsx")
    ap.add_argument("--case-counts", default="agent_space_xpu/case_counts.tsv")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    case_counts = load_case_counts(Path(args.case_counts))
    print(f"Loaded {len(case_counts)} (file,class) case counts")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(hdr)}

    if "case_count" in idx:
        cc_col = idx["case_count"]
    else:
        cc_col = ws.max_column + 1
        c = ws.cell(1, cc_col, "case_count")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    matched = 0
    rows = []
    for r in range(2, ws.max_row + 1):
        file = str(ws.cell(r, idx["file"]).value or "")
        cls = str(ws.cell(r, idx["class"]).value or "")
        cat = str(ws.cell(r, idx["category"]).value or "")
        xpu = str(ws.cell(r, idx["xpu_enabled"]).value or "")
        prio = norm_priority(ws.cell(r, idx["ut_priority"]).value)
        n = case_counts.get((file, cls))
        if n is not None:
            ws.cell(r, cc_col, n)
            matched += 1
        else:
            ws.cell(r, cc_col, "")
        rows.append({"cat": cat, "xpu": xpu, "prio": prio, "cases": n or 0,
                     "has_count": n is not None})
    print(f"case_count matched {matched}/{len(rows)} rows")

    def analyze(subset, weight):
        cat_w = defaultdict(int)
        total = 0
        agn_xpu = defaultdict(int)
        for row in subset:
            w = 1 if weight == "class" else row["cases"]
            cat_w[row["cat"]] += w
            total += w
            if row["cat"] == "device_agnostic":
                key = "True" if row["xpu"] == "True" else "False"
                agn_xpu[key] += w
        agn_total = cat_w.get("device_agnostic", 0)
        return cat_w, total, agn_xpu, agn_total

    all_rows = rows
    subset_rows = [r for r in rows if r["prio"] in PRIORITY_SUBSET]

    a1 = analyze(all_rows, "class")
    a2 = analyze(subset_rows, "class")
    a4a = analyze(all_rows, "case")
    a4b = analyze(subset_rows, "case")

    if "summary" in wb.sheetnames:
        del wb["summary"]
    sm = wb.create_sheet("summary")

    title_fill = PatternFill("solid", fgColor="4472C4")
    title_font = Font(bold=True, color="FFFFFF")
    sub_font = Font(bold=True)

    r = 1

    def write_block(header, data, total, agn_xpu, agn_total, weight_label):
        nonlocal r
        c = sm.cell(r, 1, header)
        c.font = title_font
        c.fill = title_fill
        sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        for col, h in enumerate(["category", f"count ({weight_label})", "percent", ""], 1):
            cell = sm.cell(r, col, h)
            cell.font = sub_font
        r += 1
        for cat in CATEGORIES:
            sm.cell(r, 1, cat)
            sm.cell(r, 2, data.get(cat, 0))
            sm.cell(r, 3, f"{pct(data.get(cat, 0), total)}%")
            r += 1
        sm.cell(r, 1, "TOTAL").font = sub_font
        sm.cell(r, 2, total).font = sub_font
        sm.cell(r, 3, "100.0%").font = sub_font
        r += 2
        sm.cell(r, 1, "xpu_enabled within device_agnostic").font = sub_font
        r += 1
        for k in ["True", "False"]:
            sm.cell(r, 1, k)
            sm.cell(r, 2, agn_xpu.get(k, 0))
            sm.cell(r, 3, f"{pct(agn_xpu.get(k, 0), agn_total)}%")
            r += 1
        sm.cell(r, 1, "device_agnostic TOTAL").font = sub_font
        sm.cell(r, 2, agn_total).font = sub_font
        r += 2

    write_block("#1 By CLASS count - all test classes", *a1, "classes")
    write_block("#2 By CLASS count - ut_priority in {P0,P1,P2,P3,blank}", *a2, "classes")
    write_block("#4a By CASE count - all test classes", *a4a, "cases")
    write_block("#4b By CASE count - ut_priority in {P0,P1,P2,P3,blank}", *a4b, "cases")

    note = sm.cell(r, 1, "Note: case_count from `pytest --collect-only` per file, "
                   "mapped to generic class names (device suffixes stripped). "
                   "Files failing collection have blank case_count and contribute 0 to case totals.")
    note.font = Font(italic=True, size=9)

    sm.column_dimensions["A"].width = 52
    sm.column_dimensions["B"].width = 18
    sm.column_dimensions["C"].width = 12

    wb.save(xlsx)

    print("\n===== SUMMARY =====")
    for label, (cat_w, total, agn_xpu, agn_total) in [
        ("#1 class/all", a1), ("#2 class/subset", a2),
        ("#4a case/all", a4a), ("#4b case/subset", a4b),
    ]:
        print(f"\n{label}  (total={total})")
        for cat in CATEGORIES:
            print(f"  {cat:20s} {cat_w.get(cat,0):>8} {pct(cat_w.get(cat,0),total):>6}%")
        print(f"  device_agnostic xpu_enabled True={agn_xpu.get('True',0)} "
              f"({pct(agn_xpu.get('True',0),agn_total)}%) "
              f"False={agn_xpu.get('False',0)} ({pct(agn_xpu.get('False',0),agn_total)}%)")
    print(f"\nSaved to {xlsx}")


if __name__ == "__main__":
    main()
