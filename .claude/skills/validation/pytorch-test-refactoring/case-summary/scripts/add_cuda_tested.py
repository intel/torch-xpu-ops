#!/usr/bin/env python3
"""Add a CUDA_TESTED column to the merged classification xlsx.

For each row in the target sheet, looks up (file, class) - with "test/"
prepended to the xlsx `file` value - against the (file, class) pairs found in
cuda_classes_uniq.csv (e.g. "test/backends/xeon/test_launch.py,TestTorchrun").
A row also matches if the CSV class has a "CUDA" suffix relative to the xlsx
class (e.g. xlsx class "TestBasics" matches CSV
"test/test_maskedtensor.py,TestBasicsCUDA"). Sets CUDA_TESTED to True if
either pair exists in the CSV, otherwise False.
"""

import argparse
import csv
from pathlib import Path

import openpyxl


def load_cuda_pairs(csv_path: Path) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    with open(csv_path, newline="") as f:
        for row in csv.reader(f):
            if len(row) < 2:
                continue
            pairs.add((row[0].strip(), row[1].strip()))
    return pairs


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx",
        default="results/test_class_classification_merged.xlsx",
    )
    ap.add_argument(
        "--cuda-csv",
        default="results/cuda_classes_uniq.csv",
    )
    ap.add_argument("--sheet", default="merged")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    cuda_pairs = load_cuda_pairs(Path(args.cuda_csv))
    print(f"Loaded {len(cuda_pairs)} (file,class) pairs from {args.cuda_csv}")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb[args.sheet]

    header = [c.value for c in ws[1]]
    if "CUDA_TESTED" in header:
        col_idx = header.index("CUDA_TESTED") + 1
    else:
        col_idx = len(header) + 1
        ws.cell(row=1, column=col_idx, value="CUDA_TESTED")

    file_col = header.index("file") + 1
    class_col = header.index("class") + 1

    matched = 0
    for row in range(2, ws.max_row + 1):
        file_val = ws.cell(row=row, column=file_col).value
        class_val = ws.cell(row=row, column=class_col).value
        csv_file = "test/" + file_val if file_val else file_val
        tested = (csv_file, class_val) in cuda_pairs
        if not tested and class_val:
            tested = (csv_file, class_val + "CUDA") in cuda_pairs
        matched += tested
        ws.cell(row=row, column=col_idx, value=tested)

    wb.save(xlsx)
    print(f"CUDA_TESTED=True for {matched} of {ws.max_row - 1} rows")
    print(f"Saved {xlsx}")


if __name__ == "__main__":
    main()
