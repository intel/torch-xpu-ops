#!/usr/bin/env python3
"""Stage 2: merge tracker + UT-upstream-status columns into the base
classification table, producing test_class_classification_merged.xlsx.

Reads:
  --tracker    "Device-Generic Refactoring Test Class Tracker.xlsx" (multi-sheet)
  --ut-status  "UT_Upstream_Status.xlsx" (sheet: test_all_classes)
  --src        test_class_classification.xlsx  (stage 1 output)
Writes:
  --out        test_class_classification_merged.xlsx

Tracker is left-joined by (basename(file), class); UT status by full relative
path first, then basename. PR URLs found in tracker notes are checked for merge
status via `gh` (best-effort; failures count as not-merged). Pass --no-pr-check
to skip network calls entirely.
"""

import argparse
import json
import os
import re
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract_class_name(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    m = re.search(r'"(\w+)"\s*\)\s*$', s)
    return m.group(1) if m else s


def extract_filename(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    m = re.search(r"/([^/]+\.py)(?:#|$)", s)
    if m:
        return m.group(1)
    return Path(s).name if s else ""


def is_section_header(class_name: str) -> bool:
    return (
        not class_name
        or "\u203a" in class_name  # >
        or "\u2014" in class_name  # em dash
        or class_name == "Test Class"
    )


def coerce_bool(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return str(val)
    return str(val).strip()


PR_RE = re.compile(r"https://github\.com/[^/\s]+/[^/\s]+/pull/\d+")


def _combine_notes(*parts: str) -> str:
    seen: list[str] = []
    for p in parts:
        p = p.strip()
        if p and p not in seen:
            seen.append(p)
    return "  ".join(seen)


def extract_pr_urls(text: str) -> list[str]:
    return PR_RE.findall(text)


def batch_check_pr_merged(urls: list[str]) -> dict[str, bool]:
    result: dict[str, bool] = {}
    for url in urls:
        m = re.search(r"github\.com/([^/]+/[^/]+)/pull/(\d+)", url)
        if not m:
            continue
        repo, pr_num = m.group(1), m.group(2)
        try:
            out = subprocess.run(
                [
                    "gh",
                    "api",
                    f"repos/{repo}/pulls/{pr_num}",
                    "--jq",
                    "{merged: .merged, labels: [.labels[].name]}",
                ],
                capture_output=True,
                text=True,
                timeout=15,
            )
            d = json.loads(out.stdout.strip())
            result[url] = d.get("merged", False) or "Merged" in d.get("labels", [])
        except Exception:
            result[url] = False
    return result


# Per-sheet column schema for the tracker workbook.
SHEET_SCHEMA = {
    "generic": dict(
        data_start=2, class_col=1, file_col=2, nr_col=7, pri_col=8, poc_col=9,
        notes_col=10,
    ),
    "profiler": dict(
        data_start=3, class_col=2, file_col=1, nr_col=4, pri_col=None, poc_col=5,
        notes_col=7, pr_col=6,
    ),
    "distributed": dict(
        data_start=3, class_col=1, file_col=2, nr_col=5, pri_col=None, poc_col=6,
        notes_col=None,
    ),
    "dynamo": dict(
        data_start=3, class_col=1, file_col=2, nr_col=7, pri_col=None, poc_col=8,
        notes_col=10,
    ),
    "cuda-specific": dict(
        data_start=3, class_col=1, file_col=2, nr_col=7, pri_col=None, poc_col=None,
        notes_col=5,
    ),
}


def extract_tracker_rows(wb) -> dict[tuple[str, str], dict]:
    lookup: dict[tuple[str, str], dict] = {}
    for sheet_name, schema in SHEET_SCHEMA.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [warn] sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        hits = 0
        for r in range(schema["data_start"], ws.max_row + 1):
            raw_class = ws.cell(r, schema["class_col"]).value
            class_name = extract_class_name(raw_class)
            if is_section_header(class_name):
                continue
            raw_file = (
                ws.cell(r, schema["file_col"]).value if schema["file_col"] else raw_class
            )
            filename = extract_filename(raw_file)
            if not filename and schema["file_col"] == schema["class_col"]:
                filename = extract_filename(raw_class)
            if not class_name:
                continue
            key = (filename, class_name)

            def get(col):
                return coerce_bool(ws.cell(r, col).value) if col else ""

            row = {
                "needs_refactor": get(schema["nr_col"]),
                "priority": get(schema["pri_col"]),
                "refactor_poc": get(schema["poc_col"]),
                "notes": _combine_notes(
                    get(schema["notes_col"]), get(schema.get("pr_col"))
                ),
                "tracker_sheet": sheet_name,
            }
            if key not in lookup:
                lookup[key] = row
                hits += 1
        print(f"  {sheet_name}: {hits} class rows extracted")
    return lookup


def extract_ut_status_rows(wb) -> dict[tuple[str, str], dict]:
    lookup: dict[tuple[str, str], dict] = {}
    sheet_name = "test_all_classes"
    if sheet_name not in wb.sheetnames:
        print(f"  [warn] sheet '{sheet_name}' not found in UT status xlsx")
        return lookup
    ws = wb[sheet_name]
    hits = 0
    for r in range(2, ws.max_row + 1):
        file_val = ws.cell(r, 1).value
        class_val = ws.cell(r, 2).value
        if not file_val or not class_val:
            continue
        key = (str(file_val).strip(), str(class_val).strip())
        if key not in lookup:
            lookup[key] = {
                "owner": str(ws.cell(r, 6).value or "").strip(),
                "ut_priority": str(ws.cell(r, 8).value or "").strip(),
                "status": str(ws.cell(r, 9).value or "").strip(),
                "q2": str(ws.cell(r, 10).value or "").strip(),
            }
            hits += 1
    print(f"  test_all_classes: {hits} class rows extracted")
    return lookup


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    # scripts/ -> gen_test_class_table/ -> pytorch-test-refactoring/ (holds the xlsx)
    ptr_dir = Path(__file__).resolve().parents[2]
    default_tracker = ptr_dir / "Device-Generic Refactoring Test Class Tracker.xlsx"
    default_ut = ptr_dir / "UT_Upstream_Status.xlsx"
    parser.add_argument("--tracker", default=os.environ.get("TRACKER_XLSX", str(default_tracker)))
    parser.add_argument("--ut-status", default=os.environ.get("UT_STATUS_XLSX", str(default_ut)))
    parser.add_argument(
        "--src",
        default=os.environ.get(
            "OUT_CLASSIFICATION_XLSX", str(Path.cwd() / "test_class_classification.xlsx")
        ),
    )
    parser.add_argument(
        "--out",
        default=os.environ.get(
            "OUT_MERGED_XLSX", str(Path.cwd() / "test_class_classification_merged.xlsx")
        ),
    )
    parser.add_argument(
        "--no-pr-check",
        action="store_true",
        help="Skip gh PR merge-status lookups (offline; Merged column left blank)",
    )
    args = parser.parse_args()

    tracker_xlsx = Path(args.tracker)
    ut_xlsx = Path(args.ut_status)
    src_xlsx = Path(args.src)
    out_xlsx = Path(args.out)

    print(f"Tracker   : {tracker_xlsx}")
    print(f"UT Status : {ut_xlsx}")
    print(f"Source    : {src_xlsx}")
    print(f"Output    : {out_xlsx}")

    for p in (tracker_xlsx, ut_xlsx, src_xlsx):
        if not p.exists():
            print(f"ERROR: file not found: {p}", file=sys.stderr)
            sys.exit(1)

    print("\nLoading tracker...")
    tracker_lookup = extract_tracker_rows(
        openpyxl.load_workbook(str(tracker_xlsx), data_only=True)
    )
    print(f"Total unique (file, class) keys from tracker: {len(tracker_lookup)}")

    print("\nLoading UT upstream status...")
    ut_lookup = extract_ut_status_rows(
        openpyxl.load_workbook(str(ut_xlsx), data_only=True)
    )
    print(f"Total unique (file, class) keys from UT status: {len(ut_lookup)}")

    ut_lookup_basename: dict[tuple[str, str], dict] = {}
    for (file_rel, cls), v in ut_lookup.items():
        ut_lookup_basename.setdefault((Path(file_rel).name, cls), v)

    print("\nLoading classification xlsx...")
    wb_src = openpyxl.load_workbook(str(src_xlsx))
    ws_src = wb_src.active
    total_rows = ws_src.max_row - 1
    print(f"Classification rows: {total_rows}")

    all_pr_urls: set[str] = set()
    for row in tracker_lookup.values():
        all_pr_urls.update(extract_pr_urls(row["notes"]))
    if args.no_pr_check:
        print(f"\nSkipping PR merge-status check ({len(all_pr_urls)} URLs, --no-pr-check)")
        pr_merged: dict[str, bool] = {}
    else:
        print(f"\nChecking merge status for {len(all_pr_urls)} PR URLs via gh...")
        pr_merged = batch_check_pr_merged(sorted(all_pr_urls))
        merged_count = sum(1 for v in pr_merged.values() if v)
        print(f"  Merged: {merged_count}, Not merged: {len(pr_merged) - merged_count}")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "merged"

    headers = [
        "file", "class", "category", "xpu_enabled", "needs_refactor", "priority",
        "refactor_poc", "refactoring_notes", "tracker_sheet", "owner", "ut_priority",
        "status", "Q2", "Merged",
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cat_fills = {
        "device_specific": PatternFill("solid", fgColor="FFB3B3"),
        "device_agnostic": PatternFill("solid", fgColor="B3D9FF"),
        "device_unrelated": PatternFill("solid", fgColor="D9FFB3"),
    }
    tracker_matched = ut_matched = 0

    for src_row in range(2, ws_src.max_row + 1):
        file_path = str(ws_src.cell(src_row, 1).value or "")
        class_name = str(ws_src.cell(src_row, 2).value or "")
        category = str(ws_src.cell(src_row, 3).value or "")
        xpu_val = ws_src.cell(src_row, 4).value
        basename = Path(file_path).name

        tracker_row = tracker_lookup.get((basename, class_name)) or tracker_lookup.get(
            ("", class_name)
        )
        if tracker_row:
            tracker_matched += 1
        ut_row = ut_lookup.get((file_path, class_name)) or ut_lookup_basename.get(
            (basename, class_name)
        )
        if ut_row:
            ut_matched += 1

        out_row = src_row
        ws_out.cell(out_row, 1, file_path)
        ws_out.cell(out_row, 2, class_name)
        cat_cell = ws_out.cell(out_row, 3, category)
        cat_cell.fill = cat_fills.get(category, PatternFill())

        if xpu_val in ("True", True):
            c = ws_out.cell(out_row, 4, "True")
            c.fill = PatternFill("solid", fgColor="90EE90")
        elif xpu_val in ("False", False):
            c = ws_out.cell(out_row, 4, "False")
            c.fill = PatternFill("solid", fgColor="FFD580")
        else:
            ws_out.cell(out_row, 4, "")

        if tracker_row:
            ws_out.cell(out_row, 5, tracker_row["needs_refactor"])
            ws_out.cell(out_row, 6, tracker_row["priority"])
            ws_out.cell(out_row, 7, tracker_row["refactor_poc"])
            ws_out.cell(out_row, 8, tracker_row["notes"])
            ws_out.cell(out_row, 9, tracker_row["tracker_sheet"])
        else:
            for col in range(5, 10):
                ws_out.cell(out_row, col, "")

        if ut_row:
            ws_out.cell(out_row, 10, ut_row["owner"])
            ws_out.cell(out_row, 11, ut_row["ut_priority"])
            ws_out.cell(out_row, 12, ut_row["status"])
            ws_out.cell(out_row, 13, ut_row["q2"])
        else:
            for col in range(10, 14):
                ws_out.cell(out_row, col, "")

        notes_text = tracker_row["notes"] if tracker_row else ""
        pr_urls = extract_pr_urls(notes_text)
        if pr_urls and not args.no_pr_check:
            is_merged = any(pr_merged.get(u, False) for u in pr_urls)
            mc = ws_out.cell(out_row, 14, "True" if is_merged else "False")
            mc.fill = PatternFill("solid", fgColor="90EE90" if is_merged else "FFD580")
        else:
            ws_out.cell(out_row, 14, "")

    col_widths = [60, 35, 18, 12, 15, 12, 30, 60, 16, 20, 15, 15, 15, 10]
    for col, width in enumerate(col_widths, 1):
        ws_out.column_dimensions[get_column_letter(col)].width = width

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(str(out_xlsx))

    denom = total_rows or 1
    print(f"\nTracker matched : {tracker_matched}/{total_rows} ({tracker_matched * 100 // denom}%)")
    print(f"UT status matched: {ut_matched}/{total_rows} ({ut_matched * 100 // denom}%)")
    print(f"Output written to: {out_xlsx}")


if __name__ == "__main__":
    main()
