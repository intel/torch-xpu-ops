#!/usr/bin/env python3
"""Stage 1: scan every PyTorch test file recursively, extract every test class,
classify each as device_unrelated / device_agnostic / device_specific, and
report xpu_enabled for device_agnostic class targets.

Output columns: file, class, category, xpu_enabled

Usage:
    python scan_all_test_classes.py --pytorch-test <DIR> --out <XLSX>

Both flags are optional; defaults come from environment variables
PYTORCH_TEST_DIR and OUT_CLASSIFICATION_XLSX, else the well-known dev paths.
"""

import argparse
import ast
import os
import re
from pathlib import Path

# ---------------------------------------------------------------------------
# Classification patterns (kept in sync with classify-test-files skill)
# ---------------------------------------------------------------------------
DEVICE_SPECIFIC_FILENAMES = {
    "test_cuda.py",
    "test_mps.py",
    "test_xpu.py",
    "test_rocm.py",
    "test_mtia.py",
    "test_tpu.py",
    "test_lazy.py",
}

DEVICE_SPECIFIC_IMPORTS = re.compile(
    r"(?:from\s+torch\.testing\._internal\.common_(?:cuda|mps|xpu|rocm|mtia)\s+import"
    r".*\b(TEST_CUDA|TEST_MPS|TEST_XPU|TEST_ROCM|TEST_MTIA)\b"
    r"|import\s+torch\.testing\._internal\.common_(?:cuda|mps|xpu|rocm|mtia)\b)"
)
DEVICE_SPECIFIC_APIS = re.compile(r"\btorch\.(cuda|mps|xpu)\b")

DEVICE_AGNOSTIC_IMPORTS = re.compile(
    r"from\s+torch\.testing\._internal\.common_device_type\s+import"
    r".*\binstantiate_device_type_tests\b"
    r"|from\s+torch\.testing\._internal\.common_methods_invocations\s+import"
    r"|from\s+torch\.testing\._internal\.common_(?:dtype|device_type)\s+import"
)
DEVICE_AGNOSTIC_CALL = re.compile(r"\binstantiate_device_type_tests\s*\(")
DEVICE_AGNOSTIC_DECORATORS = re.compile(
    r"@(?:ops|dtypes|dtypesIfCPU|dtypesIfCUDA|dtypesIfMPS|dtypesIfXPU)\s*\("
)
DEVICE_AGNOSTIC_ACCELERATOR = re.compile(
    r"\b(TEST_PRIVATEUSE1|TEST_ACCELERATOR)\b|torch\.accelerator\."
)


def read_file(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""


def is_device_specific(filename: str, content: str) -> bool:
    if filename in DEVICE_SPECIFIC_FILENAMES:
        return True
    if DEVICE_SPECIFIC_IMPORTS.search(content):
        return True
    if DEVICE_SPECIFIC_APIS.search(content):
        return True
    return False


def is_device_agnostic(content: str) -> bool:
    if DEVICE_AGNOSTIC_IMPORTS.search(content):
        return True
    if DEVICE_AGNOSTIC_CALL.search(content):
        return True
    if DEVICE_AGNOSTIC_DECORATORS.search(content):
        return True
    if DEVICE_AGNOSTIC_ACCELERATOR.search(content):
        return True
    return False


def classify_content(filename: str, content: str) -> str:
    if is_device_specific(filename, content):
        return "device_specific"
    if is_device_agnostic(content):
        return "device_agnostic"
    return "device_unrelated"


def get_class_source(tree: ast.AST, full_src: str, class_name: str) -> str | None:
    node = next(
        (
            n
            for n in ast.walk(tree)
            if isinstance(n, ast.ClassDef) and n.name == class_name
        ),
        None,
    )
    if node is None:
        return None
    parts = [ast.get_source_segment(full_src, node) or ""]
    for n in ast.walk(tree):
        if not isinstance(n, ast.Call):
            continue
        func = n.func
        fname = func.id if isinstance(func, ast.Name) else getattr(func, "attr", None)
        if fname != "instantiate_device_type_tests":
            continue
        first = n.args[0] if n.args else None
        if isinstance(first, ast.Name) and first.id == class_name:
            parts.append(ast.get_source_segment(full_src, n) or "")
    return "\n".join(parts)


def has_allow_xpu(tree: ast.AST, full_src: str, class_name: str) -> bool:
    for n in ast.walk(tree):
        if not isinstance(n, ast.Call):
            continue
        func = n.func
        fname = func.id if isinstance(func, ast.Name) else getattr(func, "attr", None)
        if fname != "instantiate_device_type_tests":
            continue
        first = n.args[0] if n.args else None
        if not (isinstance(first, ast.Name) and first.id == class_name):
            continue
        for kw in n.keywords:
            if kw.arg == "allow_xpu" and isinstance(kw.value, ast.Constant):
                if kw.value.value is True:
                    return True
    return False


def process_file(path: Path, pytorch_test: Path) -> list[dict]:
    src = read_file(path)
    if not src.strip():
        return []

    try:
        rel = path.relative_to(pytorch_test)
    except ValueError:
        rel = path

    try:
        tree = ast.parse(src)
    except SyntaxError:
        return []
    if tree is None:
        return []

    classes = [n for n in tree.body if isinstance(n, ast.ClassDef)]

    def _is_testcase_base(node: ast.ClassDef) -> bool:
        for base in node.bases:
            base_str = ast.unparse(base) if hasattr(ast, "unparse") else ""
            if "TestCase" in base_str or "TestCase" in getattr(base, "id", ""):
                return True
        return False

    test_classes = [
        c
        for c in classes
        if c.name.startswith("Test")
        or c.name.endswith("Test")
        or c.name.endswith("Tests")
        or _is_testcase_base(c)
    ]
    if not test_classes:
        return []

    rows = []
    for cls in test_classes:
        class_src = get_class_source(tree, src, cls.name) or ""
        category = classify_content(path.name, class_src)
        xpu_enabled = None
        if category == "device_agnostic":
            xpu_enabled = has_allow_xpu(tree, src, cls.name)
        rows.append(
            {
                "file": str(rel),
                "class": cls.name,
                "category": category,
                "xpu_enabled": "" if xpu_enabled is None else xpu_enabled,
            }
        )
    return rows


def write_xlsx(all_rows: list[dict], out_xlsx: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        import csv

        csv_path = out_xlsx.with_suffix(".csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(
                f, fieldnames=["file", "class", "category", "xpu_enabled"]
            )
            writer.writeheader()
            writer.writerows(all_rows)
        print(f"openpyxl not found; CSV written to: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test_class_classification"

    headers = ["file", "class", "category", "xpu_enabled"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cat_fills = {
        "device_specific": PatternFill("solid", fgColor="FFB3B3"),
        "device_agnostic": PatternFill("solid", fgColor="B3D9FF"),
        "device_unrelated": PatternFill("solid", fgColor="D9FFB3"),
    }

    for row_idx, row in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=row["file"])
        ws.cell(row=row_idx, column=2, value=row["class"])
        cat_cell = ws.cell(row=row_idx, column=3, value=row["category"])
        cat_cell.fill = cat_fills.get(row["category"], PatternFill())
        xpu_val = row["xpu_enabled"]
        if xpu_val is True:
            c = ws.cell(row=row_idx, column=4, value="True")
            c.fill = PatternFill("solid", fgColor="90EE90")
        elif xpu_val is False:
            c = ws.cell(row=row_idx, column=4, value="False")
            c.fill = PatternFill("solid", fgColor="FFD580")
        else:
            ws.cell(row=row_idx, column=4, value="")

    for col in range(1, 5):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(str(out_xlsx))
    print(f"Excel written to: {out_xlsx}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--pytorch-test",
        default=os.environ.get(
            "PYTORCH_TEST_DIR", str(Path.home() / "daisyden/upstream/test")
        ),
        help="Path to the PyTorch checkout's test/ directory",
    )
    parser.add_argument(
        "--out",
        default=os.environ.get(
            "OUT_CLASSIFICATION_XLSX",
            str(Path.cwd() / "test_class_classification.xlsx"),
        ),
        help="Output .xlsx path for the base classification table",
    )
    args = parser.parse_args()

    pytorch_test = Path(args.pytorch_test).resolve()
    out_xlsx = Path(args.out).resolve()
    if not pytorch_test.is_dir():
        raise SystemExit(f"ERROR: pytorch test dir not found: {pytorch_test}")

    print(f"Scanning {pytorch_test} recursively...")
    py_files = sorted(
        p
        for p in pytorch_test.rglob("*.py")
        if p.name != "__init__.py" and p.is_file()
    )
    print(f"Found {len(py_files)} Python files.")

    all_rows: list[dict] = []
    for i, f in enumerate(py_files, 1):
        if i % 200 == 0:
            print(f"  {i}/{len(py_files)} files, {len(all_rows)} classes so far...")
        all_rows.extend(process_file(f, pytorch_test))

    print(f"\nTotal test classes found: {len(all_rows)}")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(all_rows, out_xlsx)


if __name__ == "__main__":
    main()
