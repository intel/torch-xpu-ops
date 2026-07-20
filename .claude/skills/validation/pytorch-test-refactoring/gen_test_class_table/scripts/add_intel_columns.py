#!/usr/bin/env python3
"""Stage 3 (optional): append Intel Status / Intel Result / Known Issue columns
to a classification workbook (base or merged).

- Intel Status : fork branch link for (file, class) pairs committed to it.
- Intel Result : local XPU test result from a phase2 summary.tsv.
- Known Issue  : tracking-issue link where a known issue was confirmed.

Every data row gets the columns; untested / uncommitted rows are left blank.
If the columns already exist, their values are cleared and rewritten.

Usage:
    python add_intel_columns.py --xlsx <XLSX> \
        [--summary phase2_logs/summary.tsv] \
        [--committed committed_pairs.tsv] \
        [--branch-link URL] [--known-issues known_issues.tsv]

Inputs are all optional. Any missing input simply leaves that column blank.

File formats:
  summary.tsv     : TSV with header incl. columns file, class, gate, xpu_passed,
                    xpu_failed, xpu_skipped.
  committed.tsv   : 2 columns per line: <file_relpath>\\t<class>
  known_issues.tsv: 3 columns per line: <file_relpath>\\t<class>\\t<issue_text>
"""

import argparse
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GATE_TO_RESULT = {
    "PASS": "PASS",
    "TIMEOUT": "TIMEOUT (large suite, XPU rows ran)",
    "FAIL-unexplained": "FAIL",
    "FAIL-no-xpu-rows": "no XPU rows (allow_xpu no effect)",
}


def load_summary(path: Path | None) -> dict[tuple[str, str], dict]:
    out: dict[tuple[str, str], dict] = {}
    if not path or not path.exists():
        return out
    for r in csv.DictReader(open(path), delimiter="\t"):
        out[(r["file"], r["class"])] = r
    return out


def load_committed(path: Path | None) -> set[tuple[str, str]]:
    out: set[tuple[str, str]] = set()
    if not path or not path.exists():
        return out
    for line in path.read_text().splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) >= 2:
            out.add((parts[0], parts[1]))
    return out


def load_known_issues(path: Path | None) -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    if not path or not path.exists():
        return out
    for line in path.read_text().splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) >= 3:
            out[(parts[0], parts[1])] = parts[2]
    return out


def result_text(row: dict) -> str:
    gate = row["gate"]
    base = GATE_TO_RESULT.get(gate, gate)
    if gate in ("PASS", "TIMEOUT", "FAIL-unexplained"):
        return (
            f"{base} (passed={row.get('xpu_passed', '')}, "
            f"failed={row.get('xpu_failed', '')}, skipped={row.get('xpu_skipped', '')})"
        )
    return base


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Workbook to annotate in place")
    parser.add_argument("--summary", default=None)
    parser.add_argument("--committed", default=None)
    parser.add_argument("--known-issues", default=None)
    parser.add_argument(
        "--branch-link",
        default="https://github.com/daisyden/pytorch/tree/daisyden/allow_xpu",
    )
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    summary = load_summary(Path(args.summary) if args.summary else None)
    committed = load_committed(Path(args.committed) if args.committed else None)
    known = load_known_issues(Path(args.known_issues) if args.known_issues else None)

    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active

    headers = ["Intel Status", "Intel Result", "Known Issue"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    existing = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    if all(h in existing for h in headers):
        status_col, result_col, issue_col = (existing[h] for h in headers)
        for c in (status_col, result_col, issue_col):
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = PatternFill()
    else:
        ncol = ws.max_column
        status_col, result_col, issue_col = ncol + 1, ncol + 2, ncol + 3

    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=(status_col, result_col, issue_col)[i], value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    committed_fill = PatternFill("solid", fgColor="C6EFCE")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    other_fill = PatternFill("solid", fgColor="FFEB9C")
    issue_fill = PatternFill("solid", fgColor="D9D2E9")

    n_status = n_result = n_issue = 0
    for r in range(2, ws.max_row + 1):
        fpath = str(ws.cell(r, 1).value or "")
        cls = str(ws.cell(r, 2).value or "")
        key = (fpath, cls)

        if key in committed:
            sc = ws.cell(r, status_col, args.branch_link)
            sc.fill = committed_fill
            n_status += 1

        srow = summary.get(key)
        if srow:
            rc = ws.cell(r, result_col, result_text(srow))
            gate = srow["gate"]
            rc.fill = (
                pass_fill
                if gate in ("PASS", "TIMEOUT")
                else (fail_fill if gate == "FAIL-unexplained" else other_fill)
            )
            n_result += 1

        if key in known:
            ic = ws.cell(r, issue_col, known[key])
            ic.fill = issue_fill
            n_issue += 1

    ws.column_dimensions[get_column_letter(status_col)].width = 55
    ws.column_dimensions[get_column_letter(result_col)].width = 48
    ws.column_dimensions[get_column_letter(issue_col)].width = 48
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx)
    print(f"Updated {xlsx}")
    print(f"  Intel Status set: {n_status}")
    print(f"  Intel Result set: {n_result}")
    print(f"  Known Issue set : {n_issue}")


if __name__ == "__main__":
    main()
