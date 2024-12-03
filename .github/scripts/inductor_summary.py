import argparse

import pandas as pd
from scipy.stats import gmean
from styleframe import StyleFrame, Styler, utils
import numpy as np
from openpyxl import Workbook

parser = argparse.ArgumentParser(description="Generate report")
parser.add_argument('-s', '--suite', default=["huggingface"], nargs='*', type=str, help='model suite name')
parser.add_argument('-p', '--precision', default=["amp_fp16", "float32"], nargs='*', type=str, help='precision')
parser.add_argument('-r', '--reference', type=str, help='reference log files')
parser.add_argument('-m', '--mode', default=["inference", "training"], nargs='*', type=str, help='mode name')
parser.add_argument('-sc', '--scenario', default=["performance"], nargs='*', type=str, help='Test scenario set')
args = parser.parse_args()

passrate_values = {}
geomean_values = {}

new_performance_regression=pd.DataFrame()

failure_style = Styler(bg_color='#FF0000', font_color=utils.colors.black)
regression_style = Styler(bg_color='#F0E68C', font_color=utils.colors.red)
improve_style = Styler(bg_color='#00FF00', font_color=utils.colors.black)

# refer https://github.com/pytorch/pytorch/blob/main/benchmarks/dynamo/runner.py#L757-L778


def percentage(part, whole, decimals=2):
    if whole == 0:
        return 0
    return round(100 * float(part) / float(whole), decimals)


def get_passing_entries(df, column_name, scenario):
    if scenario == 'performance':
        return df[column_name][df[column_name] > 0]
    else:
        return df[df[column_name].notnull()]


def caculate_geomean(df, column_name, scenario):
    cleaned_df = get_passing_entries(df, column_name, scenario).clip(1)
    if cleaned_df.empty:
        return "0.0x"
    return f"{gmean(cleaned_df):.2f}x"


def caculate_passrate(df, key_word, scenario):
    total = len(df.index)
    if scenario == 'performance':
        passing = df[df[key_word] > 0.0][key_word].count()
    else:
        df = get_passing_entries(df, key_word, scenario)
        passing = df[df[key_word].fillna('').str.contains('pass')][key_word].count()
    perc = int(percentage(passing, total, decimals=0))
    return f"{perc}%, {passing}/{total}"

def get_perf_csv(scenario, precision, mode, suite):
    target_path = 'inductor_log/' + suite + '/' + precision + '/inductor_' + suite + '_' + precision + '_' + mode + '_xpu_' + scenario + '.csv'
    target_ori_data = pd.read_csv(target_path, header=0, encoding='utf-8')
    target_data = target_ori_data.copy()
    target_data.sort_values(by=['name'])
    
    if args.reference is not None:
        reference_file_path = args.reference + '/inductor_log/' + suite + '/' + precision + '/inductor_' + suite + '_' + precision + '_' + mode + '_xpu_' + scenario + '.csv'
        reference_ori_data = pd.read_csv(reference_file_path, header=0, encoding='utf-8')
        reference_data = reference_ori_data.copy()
        reference_data.sort_values(by=['name'])
        data = pd.merge(target_data,reference_data,on=['name'],how= 'outer')
        return data
    else:
        return target_data

def process(input, scenario, precision, mode):
    global geomean_values, passrate_values
    processed_data = pd.DataFrame()
    if input is not None and scenario == 'performance':
        if args.reference is None:
            data_new = input[['name', 'batch_size', 'speedup', 'abs_latency','compilation_latency']].rename(columns={'name': 'name', 'batch_size': 'batch_size', 'speedup': 'speedup', "abs_latency": 'inductor', "compilation_latency": 'compilation_latency'})
            data_new['inductor'] = data_new['inductor'].apply(pd.to_numeric, errors='coerce').div(1000)
            data_new['speedup'] = data_new['speedup'].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            data_new['eager'] = data_new['speedup'] * data_new['inductor']
            geomean_values['target_' + str(precision) + '_' + str(mode)] = caculate_geomean(data_new, 'speedup', scenario)
            passrate_values['target_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_new, 'inductor', scenario)
            processed_data = StyleFrame({'name': list(data_new['name']),
                            'batch_size': list(data_new['batch_size']),
                            'speedup': list(data_new['speedup']),
                            'inductor': list(data_new['inductor']),
                            'eager': list(data_new['eager']),
                            'compilation_latency': list(data_new['compilation_latency'])})
            processed_data.set_column_width(1, 10)
            processed_data.set_column_width(2, 18)
            processed_data.set_column_width(3, 18)
            processed_data.set_column_width(4, 18)
            processed_data.set_column_width(5, 15)
            processed_data.set_column_width(6, 20)
            processed_data.set_row_height(rows=processed_data.row_indexes, height=15)
        else:
            data_new=input[['name','batch_size_x','speedup_x','abs_latency_x','compilation_latency_x']].rename(columns={'name':'name','batch_size_x':'batch_size_new','speedup_x':'speed_up_new',"abs_latency_x":'inductor_new',"compilation_latency_x":'compilation_latency_new'})
            data_new['inductor_new']=data_new['inductor_new'].astype(float).div(1000)
            data_new['speed_up_new']=data_new['speed_up_new'].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            data_new['eager_new'] = data_new['speed_up_new'] * data_new['inductor_new']
            geomean_values['target_' + str(precision) + '_' + str(mode)] = caculate_geomean(data_new, 'speed_up_new', scenario)
            passrate_values['target_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_new, 'inductor_new', scenario)        
            data_old=input[['batch_size_y','speedup_y','abs_latency_y','compilation_latency_y']].rename(columns={'batch_size_y':'batch_size_old','speedup_y':'speed_up_old',"abs_latency_y":'inductor_old',"compilation_latency_y":'compilation_latency_old'}) 
            data_old['inductor_old']=data_old['inductor_old'].astype(float).div(1000)
            data_old['speed_up_old']=data_old['speed_up_old'].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            data_old['eager_old'] = data_old['speed_up_old'] * data_old['inductor_old']
            input['speedup_x']=input['speedup_x'].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            input['speedup_y']=input['speedup_y'].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            geomean_values['reference_' + str(precision) + '_' + str(mode)] = caculate_geomean(data_old, 'speed_up_old', scenario)
            passrate_values['reference_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_old, 'inductor_old', scenario)        
            data_ratio= pd.DataFrame(round(input['speedup_x'] / input['speedup_y'],2),columns=['Ratio Speedup(New/old)'])
            data_ratio['Eager Ratio(old/new)'] = pd.DataFrame(round(data_old['eager_old'] / data_new['eager_new'],2))
            data_ratio['Inductor Ratio(old/new)'] = pd.DataFrame(round(data_old['inductor_old'] / data_new['inductor_new'],2))
            data_ratio['Compilation_latency_Ratio(old/new)'] = pd.DataFrame(round(data_old['compilation_latency_old'] / data_new['compilation_latency_new'],2))
                
            combined_data = pd.DataFrame({
                'name': list(data_new['name']),
                'batch_size_new': list(data_new['batch_size_new']),
                'speed_up_new': list(data_new['speed_up_new']),
                'inductor_new': list(data_new['inductor_new']),
                'eager_new': list(data_new['eager_new']),
                'compilation_latency_new': list(data_new['compilation_latency_new']),
                'batch_size_old': list(data_old['batch_size_old']),
                'speed_up_old': list(data_old['speed_up_old']),
                'inductor_old': list(data_old['inductor_old']),
                'eager_old': list(data_old['eager_old']),
                'compilation_latency_old': list(data_old['compilation_latency_old']),
                'Ratio Speedup(New/old)': list(data_ratio['Ratio Speedup(New/old)']),
                'Eager Ratio(old/new)': list(data_ratio['Eager Ratio(old/new)']),
                'Inductor Ratio(old/new)': list(data_ratio['Inductor Ratio(old/new)']),
                'Compilation_latency_Ratio(old/new)': list(data_ratio['Compilation_latency_Ratio(old/new)'])
                }) 
            processed_data = StyleFrame(combined_data)
            processed_data.set_column_width(1, 10)
            processed_data.set_column_width(2, 18) 
            processed_data.set_column_width(3, 18) 
            processed_data.set_column_width(4, 18)
            processed_data.set_column_width(5, 15)
            processed_data.set_column_width(6, 20)
            processed_data.set_column_width(7, 18)
            processed_data.set_column_width(8, 18) 
            processed_data.set_column_width(9, 18) 
            processed_data.set_column_width(10, 15)
            processed_data.set_column_width(11, 20)
            processed_data.set_column_width(12, 28)
            processed_data.set_column_width(13, 28) 
            processed_data.set_column_width(14, 28)
            processed_data.set_column_width(15, 32)
            processed_data.apply_style_by_indexes(indexes_to_style=processed_data[processed_data['batch_size_new'] == 0], styler_obj=failure_style)
            processed_data.apply_style_by_indexes(indexes_to_style=processed_data[(processed_data['Inductor Ratio(old/new)'] > 0) & (processed_data['Inductor Ratio(old/new)'] < 0.9)],styler_obj=regression_style)
            global new_performance_regression
            regression = processed_data.loc[(processed_data['Inductor Ratio(old/new)'] > 0) & (processed_data['Inductor Ratio(old/new)'] < 0.9)]
            regression = regression.copy()
            regression.loc[0] = list(regression.shape[1]*'*')
            new_performance_regression = pd.concat([new_performance_regression,regression])
            processed_data.apply_style_by_indexes(indexes_to_style=processed_data[processed_data['Inductor Ratio(old/new)'] > 1.1],styler_obj=improve_style)
            processed_data.set_row_height(rows=processed_data.row_indexes, height=15)        
    if input is not None and scenario == 'accuracy':
        if args.reference is None:
            data_new = input[['name', 'batch_size', 'accuracy']].rename(columns={'name': 'name', 'batch_size': 'batch_size', 'accuracy': 'accuracy'})
            passrate_values['target_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_new, 'accuracy', scenario)
            processed_data = data_new
            processed_data = StyleFrame({'name': list(data_new['name']),
                            'batch_size': list(data_new['batch_size']),
                             'accuracy': list(data_new['accuracy'])})
            processed_data.set_column_width(1, 10)
            processed_data.set_column_width(2, 18)
            processed_data.set_column_width(3, 18)
            processed_data.set_row_height(rows=processed_data.row_indexes, height=15)
        else:
            data_new=input[['name','batch_size_x','accuracy_x']].rename(columns={'name':'name','batch_size_x':'batch_size_new','accuracy_x':'accuracy_new'})
            passrate_values['target_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_new, 'accuracy_new', scenario)        
            data_old=input[['batch_size_y','accuracy_y']].rename(columns={'batch_size_y':'batch_size_old','accuracy_y':'accuracy_old'}) 
            passrate_values['reference_' + str(precision) + '_' + str(mode)] = caculate_passrate(data_old, 'accuracy_old', scenario)
            data_comp = pd.DataFrame((data_new['accuracy_new'] != 'pass') & (data_old['accuracy_old'] == 'pass'),columns=['Accuracy regression'])
            combined_data = pd.DataFrame({
                'name': list(data_new['name']),
                'batch_size_new': list(data_new['batch_size_new']),
                'accuracy_new': list(data_new['accuracy_new']),
                'batch_size_old': list(data_old['batch_size_old']),
                'accuracy_old': list(data_old['accuracy_old']),
                'Accuracy regression': list(data_comp['Accuracy regression'])
                })
            processed_data = StyleFrame(combined_data)
            processed_data.set_column_width(1, 10)
            processed_data.set_column_width(2, 18) 
            processed_data.set_column_width(3, 18) 
            processed_data.set_column_width(4, 18)
            processed_data.set_column_width(5, 15)
            processed_data.set_column_width(6, 20)
            processed_data.apply_style_by_indexes(indexes_to_style=processed_data[(processed_data['Accuracy regression'] == 'regression')],styler_obj=regression_style)
            processed_data.set_row_height(rows=processed_data.row_indexes, height=15)        
    return processed_data

def update_details(scenario, precision, mode, suite, excel):
    if scenario == 'performance':
        h = {"A": 'Model suite', "B": '', "C": "target", "D": '', "E": '', "F": '', "G": '',"H": args.reference, "I": '', "J": '',"K": '',"L":'',"M": 'Result Comp',"N": '',"O": '',"P":''}
        if args.reference is None:
            h = {"A": 'Model suite', "B": '', "C": "target", "D": '', "E": '', "F": '', "G": ''}
    else:
        h = {"A": 'Model suite', "B": '', "C": "target", "D": '', "E": args.reference, "F": '', "G": 'Result Comp'}
        if args.reference is None:
            h = {"A": 'Model suite', "B": '', "C": "target", "D": ''}
    head = StyleFrame(pd.DataFrame(h, index=[0]))
    head.set_column_width(1, 15)
    head.set_row_height(rows=[1], height=15)
    head.to_excel(excel_writer=excel, sheet_name=suite + '_' + precision + '_' + mode[0:3] + '_' + scenario[0:3], index=False, startrow=0, header=False)
    target_raw_data = get_perf_csv(scenario, precision, mode, suite)
    target_data = process(target_raw_data, scenario, precision, mode)
    target_data.to_excel(excel_writer=excel, sheet_name=suite + '_' + precision + '_' + mode[0:3] + '_' + scenario[0:3], index=False, startrow=1, startcol=1)

def update_summary(excel, scenario, suite):
    if scenario == 'performance':
        data = {
            'Test Secnario': ['AMP_BF16 Inference', ' ', 'AMP_BF16 Training', ' ', 'AMP_FP16 Inference', ' ', 'AMP_FP16 Training', ' ', 'BF16 Inference', ' ', 'BF16 Training', ' ', 'FP16 Inference', ' ', 'FP16 Training', ' ', 'FP32 Inference', ' ', 'FP32 Training', ' '],
            'Comp Item': ['Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup', 'Pass Rate', 'Geomean Speedup'],
            'Date': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'Compiler': ['inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor'],
            'torchbench': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'huggingface': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'timm_models ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'ref_torchbench ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'refer_huggingface ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
            'refer_timm_models ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        }
        summary = pd.DataFrame(data)

        if suite == 'huggingface':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[0:1, 5:6] = passrate_values['target_amp_bf16_inference']
                    summary.iloc[1:2, 5:6] = geomean_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[2:3, 5:6] = passrate_values['target_amp_bf16_training']
                    summary.iloc[3:4, 5:6] = geomean_values['target_amp_bf16_training']
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 5:6] = passrate_values['target_amp_fp16_inference']
                    summary.iloc[5:6, 5:6] = geomean_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[6:7, 5:6] = passrate_values['target_amp_fp16_training']
                    summary.iloc[7:8, 5:6] = geomean_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 5:6] = passrate_values['target_bfloat16_inference']
                    summary.iloc[9:10, 5:6] = geomean_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[10:11, 5:6] = passrate_values['target_bfloat16_training']
                    summary.iloc[11:12, 5:6] = geomean_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[12:13, 5:6] = passrate_values['target_float16_inference']
                    summary.iloc[13:14, 5:6] = geomean_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[14:15, 5:6] = passrate_values['target_float16_training']
                    summary.iloc[15:16, 5:6] = geomean_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[16:17, 5:6] = passrate_values['target_float32_inference']
                    summary.iloc[17:18, 5:6] = geomean_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[18:19, 5:6] = passrate_values['target_float32_training']
                    summary.iloc[19:20, 5:6] = geomean_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[0:1, 8:9] = passrate_values['reference_amp_bf16_inference']
                        summary.iloc[1:2, 8:9] = geomean_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[2:3, 8:9] = passrate_values['reference_amp_bf16_training']
                        summary.iloc[3:4, 8:9] = geomean_values['reference_amp_bf16_training']
                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 8:9] = passrate_values['reference_amp_fp16_inference']
                        summary.iloc[5:6, 8:9] = geomean_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[6:7, 8:9] = passrate_values['reference_amp_fp16_training']
                        summary.iloc[7:8, 8:9] = geomean_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 8:9] = passrate_values['reference_bfloat16_inference']
                        summary.iloc[9:10, 8:9] = geomean_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[10:11, 8:9] = passrate_values['reference_bfloat16_training']
                        summary.iloc[11:12, 8:9] = geomean_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[12:13, 8:9] = passrate_values['reference_float16_inference']
                        summary.iloc[13:14, 8:9] = geomean_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[14:15, 8:9] = passrate_values['reference_float16_training']
                        summary.iloc[15:16, 8:9] = geomean_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[16:17, 8:9] = passrate_values['reference_float32_inference']
                        summary.iloc[17:18, 8:9] = geomean_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[18:19, 8:9] = passrate_values['reference_float32_training']
                        summary.iloc[19:20, 8:9] = geomean_values['reference_float32_training']

        if suite == 'timm_models':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[0:1, 6:7] = passrate_values['target_amp_bf16_inference']
                    summary.iloc[1:2, 6:7] = geomean_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[2:3, 6:7] = passrate_values['target_amp_bf16_training']
                    summary.iloc[3:4, 6:7] = geomean_values['target_amp_bf16_training']
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 6:7] = passrate_values['target_amp_fp16_inference']
                    summary.iloc[5:6, 6:7] = geomean_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[6:7, 6:7] = passrate_values['target_amp_fp16_training']
                    summary.iloc[7:8, 6:7] = geomean_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 6:7] = passrate_values['target_bfloat16_inference']
                    summary.iloc[9:10, 6:7] = geomean_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[10:11, 6:7] = passrate_values['target_bfloat16_training']
                    summary.iloc[11:12, 6:7] = geomean_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[12:13, 6:7] = passrate_values['target_float16_inference']
                    summary.iloc[13:14, 6:7] = geomean_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[14:15, 6:7] = passrate_values['target_float16_training']
                    summary.iloc[15:16, 6:7] = geomean_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[16:17, 6:7] = passrate_values['target_float32_inference']
                    summary.iloc[17:18, 6:7] = geomean_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[18:19, 6:7] = passrate_values['target_float32_training']
                    summary.iloc[19:20, 6:7] = geomean_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[0:1, 9:10] = passrate_values['reference_amp_bf16_inference']
                        summary.iloc[1:2, 9:10] = geomean_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[2:3, 9:10] = passrate_values['reference_amp_bf16_training']
                        summary.iloc[3:4, 9:10] = geomean_values['reference_amp_bf16_training']

                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 9:10] = passrate_values['reference_amp_fp16_inference']
                        summary.iloc[5:6, 9:10] = geomean_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[6:7, 9:10] = passrate_values['reference_amp_fp16_training']
                        summary.iloc[7:8, 9:10] = geomean_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 9:10] = passrate_values['reference_bfloat16_inference']
                        summary.iloc[9:10, 9:10] = geomean_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[10:11, 9:10] = passrate_values['reference_bfloat16_training']
                        summary.iloc[11:12, 9:10] = geomean_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[12:13, 9:10] = passrate_values['reference_float16_inference']
                        summary.iloc[13:14, 9:10] = geomean_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[14:15, 9:10] = passrate_values['reference_float16_training']
                        summary.iloc[15:16, 9:10] = geomean_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[16:17, 9:10] = passrate_values['reference_float32_inference']
                        summary.iloc[17:18, 9:10] = geomean_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[18:19, 9:10] = passrate_values['reference_float32_training']
                        summary.iloc[19:20, 9:10] = geomean_values['reference_float32_training']
        
        if suite == 'torchbench':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[0:1, 4:5] = passrate_values['target_amp_bf16_inference']
                    summary.iloc[1:2, 4:5] = geomean_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[2:3, 4:5] = passrate_values['target_amp_bf16_training']
                    summary.iloc[3:4, 4:5] = geomean_values['target_amp_bf16_training']
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 4:5] = passrate_values['target_amp_fp16_inference']
                    summary.iloc[5:6, 4:5] = geomean_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[6:7, 4:5] = passrate_values['target_amp_fp16_training']
                    summary.iloc[7:8, 4:5] = geomean_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 4:5] = passrate_values['target_bfloat16_inference']
                    summary.iloc[9:10, 4:5] = geomean_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[10:11, 4:5] = passrate_values['target_bfloat16_training']
                    summary.iloc[11:12, 4:5] = geomean_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[12:13, 4:5] = passrate_values['target_float16_inference']
                    summary.iloc[13:14, 4:5] = geomean_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[14:15, 4:5] = passrate_values['target_float16_training']
                    summary.iloc[15:16, 4:5] = geomean_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[16:17, 4:5] = passrate_values['target_float32_inference']
                    summary.iloc[17:18, 4:5] = geomean_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[18:19, 4:5] = passrate_values['target_float32_training']
                    summary.iloc[19:20, 4:5] = geomean_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[0:1, 7:8] = passrate_values['reference_amp_bf16_inference']
                        summary.iloc[1:2, 7:8] = geomean_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[2:3, 7:8] = passrate_values['reference_amp_bf16_training']
                        summary.iloc[3:4, 7:8] = geomean_values['reference_amp_bf16_training']
                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 7:8] = passrate_values['reference_amp_fp16_inference']
                        summary.iloc[5:6, 7:8] = geomean_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[6:7, 7:8] = passrate_values['reference_amp_fp16_training']
                        summary.iloc[7:8, 7:8] = geomean_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 7:8] = passrate_values['reference_bfloat16_inference']
                        summary.iloc[9:10, 7:8] = geomean_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[10:11, 7:8] = passrate_values['reference_bfloat16_training']
                        summary.iloc[11:12, 7:8] = geomean_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[12:13, 7:8] = passrate_values['reference_float16_inference']
                        summary.iloc[13:14, 7:8] = geomean_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[14:15, 7:8] = passrate_values['reference_float16_training']
                        summary.iloc[15:16, 7:8] = geomean_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[16:17, 7:8] = passrate_values['reference_float32_inference']
                        summary.iloc[17:18, 7:8] = geomean_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[18:19, 7:8] = passrate_values['reference_float32_training']
                        summary.iloc[19:20, 7:8] = geomean_values['reference_float32_training']
            
        print(f"===================={scenario}Summary=============================")
        print(summary)
        
        sf = StyleFrame(summary)
        for i in range(1, 11):
            sf.set_column_width(i, 22)
        for j in range(1, 22):
            sf.set_row_height(j, 30)
        sf.to_excel(sheet_name=suite + '_'  + scenario + '_Summary', excel_writer=excel)

    else:
        data = {
        'Test Secnario': ['AMP_BF16 Inference', 'AMP_BF16 Training', 'AMP_FP16 Inference', 'AMP_FP16 Training', 'BF16 Inference', 'BF16 Training', 'FP16 Inference', 'FP16 Training', 'FP32 Inference', 'FP32 Training'],
        'Comp Item': ['Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate',  'Pass Rate'],
        'compiler': ['inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor', 'inductor'],
        'torchbench': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
        'huggingface': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
        'timm_models ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
        'refer_torchbench ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
        'refer_huggingface ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '],
        'refer_timm_models ': [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        }
        summary = pd.DataFrame(data)

        if suite == 'huggingface':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    # passrate
                    summary.iloc[0:1, 4:5] = passrate_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[1:2, 4:5] = passrate_values['target_amp_bf16_training']
                    
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[2:3, 4:5] = passrate_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[3:4, 4:5] = passrate_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 4:5] = passrate_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[5:6, 4:5] = passrate_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[6:7, 4:5] = passrate_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[7:8, 4:5] = passrate_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 4:5] = passrate_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[9:10, 4:5] = passrate_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                    # passrate
                        summary.iloc[0:1, 7:8] = passrate_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[1:2, 7:8] = passrate_values['reference_amp_bf16_training']
                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[2:3, 7:8] = passrate_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[3:4, 7:8] = passrate_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 7:8] = passrate_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[5:6, 7:8] = passrate_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[6:7, 7:8] = passrate_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[7:8, 7:8] = passrate_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 7:8] = passrate_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[9:10, 7:8] = passrate_values['reference_float32_training']

        if suite == 'timm_models':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    # passrate
                    summary.iloc[0:1, 5:6] = passrate_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[1:2, 5:6] = passrate_values['target_amp_bf16_training']
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[2:3, 5:6] = passrate_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[3:4, 5:6] = passrate_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 5:6] = passrate_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[5:6, 5:6] = passrate_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[6:7, 5:6] = passrate_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[7:8, 5:6] = passrate_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 5:6] = passrate_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[9:10, 5:6] = passrate_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                        # passrate
                        summary.iloc[0:1, 8:9] = passrate_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[1:2, 8:9] = passrate_values['reference_amp_bf16_training']
                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[2:3, 8:9] = passrate_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[3:4, 8:9] = passrate_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 8:9] = passrate_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[5:6, 8:9] = passrate_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[6:7, 8:9] = passrate_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[7:8, 8:9] = passrate_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 8:9] = passrate_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[9:10, 8:9] = passrate_values['reference_float32_training']

        if suite == 'torchbench':
            if 'amp_bf16' in args.precision:
                if 'inference' in args.mode:
                    # passrate
                    summary.iloc[0:1, 3:4] = passrate_values['target_amp_bf16_inference']
                if 'training' in args.mode:
                    summary.iloc[1:2, 3:4] = passrate_values['target_amp_bf16_training']
            if 'amp_fp16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[2:3, 3:4] = passrate_values['target_amp_fp16_inference']
                if 'training' in args.mode:
                    summary.iloc[3:4, 3:4] = passrate_values['target_amp_fp16_training']

            if 'bfloat16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[4:5, 3:4] = passrate_values['target_bfloat16_inference']
                if 'training' in args.mode:
                    summary.iloc[5:6, 3:4] = passrate_values['target_bfloat16_training']
            if 'float16' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[6:7, 3:4] = passrate_values['target_float16_inference']
                if 'training' in args.mode:
                    summary.iloc[7:8, 3:4] = passrate_values['target_float16_training']
            if 'float32' in args.precision:
                if 'inference' in args.mode:
                    summary.iloc[8:9, 3:4] = passrate_values['target_float32_inference']
                if 'training' in args.mode:
                    summary.iloc[9:10, 3:4] = passrate_values['target_float32_training']

            if args.reference is not None:
                if 'amp_bf16' in args.precision:
                    if 'inference' in args.mode:
                        # passrate
                        summary.iloc[0:1, 6:7] = passrate_values['reference_amp_bf16_inference']
                    if 'training' in args.mode:
                        summary.iloc[1:2, 6:7] = passrate_values['reference_amp_bf16_training']

                if 'amp_fp16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[2:3, 6:7] = passrate_values['reference_amp_fp16_inference']
                    if 'training' in args.mode:
                        summary.iloc[3:4, 6:7] = passrate_values['reference_amp_fp16_training']

                if 'bfloat16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[4:5, 6:7] = passrate_values['reference_bfloat16_inference']
                    if 'training' in args.mode:
                        summary.iloc[5:6, 6:7] = passrate_values['reference_bfloat16_training']
                if 'float16' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[6:7, 6:7] = passrate_values['reference_float16_inference']
                    if 'training' in args.mode:
                        summary.iloc[7:8, 6:7] = passrate_values['reference_float16_training']
                if 'float32' in args.precision:
                    if 'inference' in args.mode:
                        summary.iloc[8:9, 6:7] = passrate_values['reference_float32_inference']
                    if 'training' in args.mode:
                        summary.iloc[9:10, 6:7] = passrate_values['reference_float32_training']
            
        print(f"===================={scenario} Summary=============================")
        print(summary)
        
        sf = StyleFrame(summary)
        for i in range(1, 10):
            sf.set_column_width(i, 22)
        for j in range(1, 12):
            sf.set_row_height(j, 30)
        sf.to_excel(sheet_name=suite + '_'  + scenario + '_Summary', excel_writer=excel)

def summary_conclusion(scenario, excel):
    excel.book.save(excel)
    df = pd.read_excel(excel, sheet_name = None, header = None)
    #df = pd.DataFrame(excel)
    if scenario == 'performance':
        sheet_names = list(df.keys())
        sheet_names = [s for s in sheet_names if 'Summary' in s and 'performance' in s]
        sheet_names.sort()
        print(f"Merge excel as below:\n{sheet_names}")
        print("\n")
        features = [[]] * 21
        for sheet_name in sheet_names:
            df_sheet = df[sheet_name]
            df_sheet = df_sheet.values
            features = np.hstack((features, df_sheet))
        
        if len(sheet_names) == 1:
            print("sheet not merge")
        elif len(sheet_names) == 2:
            print("2 sheets merge")
            if 'huggingface' in sheet_names[0]:
                features[:, 4:5] = features[:, 14:15]
                features[:, 6:7] = features[:, 16:17]
            else:
                features[:, 4:5] = features[:, 14:15]
        else:
            print("3 sheets merge")
            features[:, 4:5] = features[:, 24:25]
            features[:, 6:7] = features[:, 16:17]

        df_concat = StyleFrame(pd.DataFrame(features).iloc[:,:10])
        for i in range(10):
            df_concat.set_column_width(i, 22)
        for j in range(1, 23):
            df_concat.set_row_height(j, 30)
        df_concat.to_excel(sheet_name='Perf_Summary', excel_writer=excel, index=False)
    else:
        sheet_names = list(df.keys())
        sheet_names = [s for s in sheet_names if 'Summary' in s and 'accuracy' in s]
        sheet_names.sort()
        print(f"Merge excel as below:\n{sheet_names}")
        print("\n")
        features = [[]] * 11
        for sheet_name in sheet_names:
            df_sheet = df[sheet_name]
            df_sheet = df_sheet.values
            features = np.hstack((features, df_sheet))
        if len(sheet_names) == 1:
            print("sheet not merge")
        elif len(sheet_names) == 2:
            print("2 sheets merge")
            if 'huggingface' in sheet_names[0]:
                features[:, 3:4] = features[:, 12:13]
                features[:, 5:6] = features[:, 14:15]
            else:
                features[:, 3:4] = features[:, 12:13]
        else:
            print("3 sheets merge")
            features[:, 3:4] = features[:, 21:22]
            features[:, 5:6] = features[:, 14:15]

        df_concat = StyleFrame(pd.DataFrame(features).iloc[:,:9])
        for i in range(10):
            df_concat.set_column_width(i, 22)
        for j in range(1, 13):
            df_concat.set_row_height(j, 30)
        df_concat.to_excel(sheet_name='Acc_Summary', excel_writer=excel, index=False)

def generate_report(excel, scenario_list, precision_list, mode_list, suite_list):
    for sc in scenario_list:
        for s in suite_list:
            for p in precision_list:
                for m in mode_list:
                    update_details(sc, p, m, s, excel)
            update_summary(excel, sc, s)


def excel_postprocess(file, scenario, precison, mode, suite):
    wb = file.book
    # Summary
    #ws = wb['Summary']
    #for i in range(2, 21, 2):
    #    ws.merge_cells(start_row=i, end_row=i + 1, start_column=1, end_column=1)
    # details
    for sc in scenario:
        for s in suite:
            for p in precison:
                for m in mode:
                    wdt = wb[s + '_' + p + '_' + m[0:3] + '_' + sc[0:3]]
                    wdt.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
                    wdt.merge_cells(start_row=1, end_row=1, start_column=3, end_column=7)
                    wdt.merge_cells(start_row=1, end_row=1, start_column=8, end_column=12)
                    wdt.merge_cells(start_row=1, end_row=1, start_column=13, end_column=16)
            wb.save(file)

        if len(scenario) == 2:
            wb.move_sheet("Perf_Summary", -(len(wb.worksheets)-1))
            wb.move_sheet("Acc_Summary", -(len(wb.worksheets)-1))
        elif len(scenario) == 1 and sc == 'accuracy':
            wb.move_sheet("Acc_Summary", -(len(wb.worksheets)-1))
        else:
            wb.move_sheet("Perf_Summary", -(len(wb.worksheets)-1))


if __name__ == '__main__':
    excel = StyleFrame.ExcelWriter('inductor_log/Inductor_E2E_Test_Report.xlsx')
    generate_report(excel, args.scenario, args.precision, args.mode, args.suite)
    for sc in args.scenario:
        summary_conclusion(sc, excel)
    excel_postprocess(excel, args.scenario, args.precision, args.mode, args.suite)
    excel.close()
