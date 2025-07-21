import re
import argparse
import openpyxl
import os
import pandas as pd

test_file_num = 0

def parse_txt_to_excel(txt_file, df_exist=None):
    with open(txt_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    pattern = re.compile(r"Running [1-9]\d* items in this shard:")
    num_pattern = re.compile(r"[1-9]\d*")
    timestamp_pattern = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")
    is_end = False
    df = pd.DataFrame(columns=["file", "totle cases", "cases"]) if df_exist is None else df_exist
    line_id = 0

    # for i, line in enumerate(lines):
    while line_id < len(lines):
        line = lines[line_id]
        line_id += 1
        item_num = None
        match = pattern.search(line)
        if match:
            run_items = []
            file_name = None
            item_num = int(num_pattern.search(match.group()).group())
            line = line.split('this shard:')[1]
            if timestamp_pattern.search(line):
                loc = timestamp_pattern.search(line).span()[0]
                line = line[:loc]
                is_end = True
            if ',' in line:
                run_items.extend([item.strip() for item in line.split(',')])
            else:
                run_items.append(line.strip())
            if not is_end:
                # for next_line in lines[i + 1:]:
                while True:
                    next_line = lines[line_id]
                    line_id += 1
                    if timestamp_pattern.search(next_line):
                        loc = timestamp_pattern.search(next_line).span()[0]
                        if loc != 0:
                            next_line = next_line[loc:]
                            if ',' in next_line:
                                run_items.extend([item.strip() for item in next_line.split(',')])
                            else:
                                run_items.append(next_line.strip())
                                
                        break
                    else:
                        if ',' in next_line:
                            run_items.extend([item.strip() for item in next_line.split(',')])
                        else:
                            run_items.append(next_line.strip())

            if file_name is None and len(run_items) > 0:
                file_name = run_items[0].split('::')[0]
            file_name_list = df[pd.notnull(df.file)].index
            global test_file_num
            test_file_num = len(file_name_list)
            test_file_num += 1
            file_name_num = df[df.file == file_name].index
            print("====file_name_num:", file_name_num)
            if file_name_num.empty:
                df = pd.concat([df, pd.DataFrame({"file": [file_name],"totle cases": [item_num], "cases": [test_file_num]})], ignore_index=True)
                df = pd.concat([df, pd.DataFrame({"cases": run_items})], ignore_index=True)
            else:
                index = file_name_list.get_loc(file_name_num[0]) + 1
                if index < len(file_name_list):
                    line_num = file_name_list[index]
                    print("file name: {} exist! line:{}, {}: line: {}".format(file_name, file_name_list[index-1], df.loc[line_num, "file"], line_num))
                    if file_name_num[0] == 19993:
                        df.to_excel("before.xlsx", index=False)
                    df = pd.concat([df.loc[:(line_num - 1)], pd.DataFrame({"cases": run_items}), df.loc[line_num:]], ignore_index=True).reset_index(drop=True)
                    if file_name_num[0] == 19993:
                        df.to_excel("after.xlsx", index=False)
                    file_name_num = df[df.file == file_name].index
                else:
                    df = pd.concat([df, pd.DataFrame({"cases": run_items})], ignore_index=True)
                org_items = df.shape[0]
                # if pd.isnull(df.loc[file_name_num[0], "file"]):
                #     import pdb;pdb.set_trace()
                print("=====all cases before dp: {}".format(org_items))
                # 去重操作，假设你想要根据列进行去重
                df = df.drop_duplicates(subset=['cases'])
                new_items = df.shape[0]
                print("=====all cases after dp: {}".format(new_items))
                try:
                    df.loc[file_name_num[0], "totle cases"] = int(df.loc[file_name_num[0], "totle cases"]) + (item_num-(org_items-new_items))
                except:
                    print("Exception got")

    print("====={}: {}".format(txt_file, df.shape))
    return df

def process_path(input_dir, output_file, sheet_title):

    if os.path.exists(output_file):
        # work_book = openpyxl.load_workbook(output_file)
        df_parsed = pd.read_excel(output_file, sheet_name=sheet_title)
        print("=====already exist: {}".format(df_parsed.shape))
        # if sheet_title in work_book.sheetnames:
        #     sheet = work_book[sheet_title]
        # else:
        #     sheet = work_book.create_sheet(sheet_title)
    else:
        df_parsed = pd.DataFrame(columns=["file", "totle cases", "cases"])
        # work_book = openpyxl.Workbook()
        # sheet = work_book.active
        # sheet.title = sheet_title

    if os.path.isfile(input_dir):
        df_parsed = parse_txt_to_excel(input_dir, df_parsed)
    elif os.path.isdir(input_dir):
        # df_parsed = pd.DataFrame(columns=["file", "totle cases", "cases"])
        for file_name in os.listdir(input_dir):
            if file_name.endswith('.txt'):
                df_parsed = parse_txt_to_excel(os.path.join(input_dir, file_name), df_parsed)
    else:
        print(f"Error: {input_dir} is neither a file nor a directory.")

    # # df = pd.concat([df, df_parsed], ignore_index=True)
    # print("=====all cases before dp: {}".format(df_parsed.shape))
    # # 去重操作，假设你想要根据列进行去重
    # df_deduplicated = df_parsed.drop_duplicates(subset=['cases'])
    # print("=====all cases after dp: {}".format(df_deduplicated.shape))

    # 将去重后的数据写回到一个新的Excel文件中
    # output_file_path = 'deduplicated_example.xlsx' # 替换为你想要保存的文件路径
    df_parsed.to_excel(output_file, sheet_name=sheet_title, index=False) # index=False表示不保存行索引

    print("已保存到文件：", output_file)

    # work_book.save(output_file)

# Example usage
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Parse cases name from log and save to Excel.')
    parser.add_argument('--input_dir', help='The input directory containing the log files.')
    parser.add_argument('-o', '--output', default='cases.xlsx', help='The output Excel file (default: cases.xlsx)')
    parser.add_argument('-s', '--sheet_name', default='cuda UT cases', help='The sheet title in the Excel file (default: cuda UT cases)')

    args = parser.parse_args()
    input_dir = args.input_dir
    output_excel_file = args.output
    sheet_title = args.sheet_name

    process_path(input_dir, output_excel_file, sheet_title)
