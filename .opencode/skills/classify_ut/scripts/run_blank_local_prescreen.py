#!/usr/bin/env python3
"""
Step 0 - Mandatory Local Pre-Screen runner.

Double-checks blank-Reason rows in the XPU skipped sheets by running the
corresponding XPU test locally and recording the outcome in a new `local_result`
column. On PASS, also writes Reason="Local Passed" + DetailReason.

This script DOES NOT modify any test code or repo files. It only invokes pytest
and records what it observed.

File resolution per row (locked policy):
    1. third_party/torch-xpu-ops/test/xpu/<basename>_xpu.py   (preferred)
    2. test/<testfile_cuda>                                    (fallback)

Node ID per row:
    classname = classname_xpu if populated else CUDA->XPU swap of classname_cuda
    name      = name_xpu      if populated else name_cuda

Sheets covered:
    - XPU skipped only Non-Inductor
    - XPU skipped only Inductor

Resume-safe: rows whose `local_result` cell is already populated are skipped.

Environment variables:
    PYTORCH_SRC                Path to pytorch source (default /home/daisyden/upstream/pytorch)
    PYTEST_PER_TEST_TIMEOUT    Per-test timeout in seconds (default 60)
    LOG_DIR                    Log directory (default /tmp/opencode/<workbook_basename>_local_verify)
    CHECKPOINT_EVERY           Save workbook every N rows (default 25)
    MAX_ROWS                   Cap rows per sheet for smoke testing (default unlimited)
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
import zlib
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


PYTORCH_SRC = Path(os.environ.get("PYTORCH_SRC", "/home/daisyden/upstream/pytorch"))
XPU_OPS_TEST_DIR = PYTORCH_SRC / "third_party" / "torch-xpu-ops" / "test" / "xpu"
PYTORCH_TEST_DIR = PYTORCH_SRC / "test"

PYTEST_TIMEOUT = int(os.environ.get("PYTEST_PER_TEST_TIMEOUT", "60"))
CHECKPOINT_EVERY = int(os.environ.get("CHECKPOINT_EVERY", "25"))
_MAX = os.environ.get("MAX_ROWS")
MAX_ROWS = int(_MAX) if _MAX else None

SHEETS = ("XPU skipped only Non-Inductor", "XPU skipped only Inductor")

BLUE = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

CONFIDENCE_PREFIX = "[Confidence: HIGH]"

CUDA_SUFFIX_RE = re.compile(r"(CUDA|Cuda|cuda)$")


def header_index(ws) -> dict[str, int]:
    return {
        cell.value: idx + 1
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
        if cell.value is not None
    }


def ensure_local_result_column(ws, cols: dict[str, int]) -> int:
    """Add `local_result` column adjacent to DetailReason if missing.

    Side effects: shifts every column to the right of DetailReason one slot to the
    right (preserving values and fills), mutates `cols` in place to reflect the
    new layout, and returns the 1-based index of `local_result`. The column-shift
    is required to keep `local_result` directly next to `DetailReason` even when
    later columns (Reason TBD, Confidence, ...) already exist.
    """
    if "local_result" in cols:
        return cols["local_result"]
    detail_col = cols["DetailReason"]
    new_col = detail_col + 1
    later_cols = sorted(
        ((name, idx) for name, idx in cols.items() if idx > detail_col),
        key=lambda x: -x[1],
    )
    for name, idx in later_cols:
        for row in ws.iter_rows(min_row=1, min_col=idx, max_col=idx):
            cell = row[0]
            tgt = ws.cell(row=cell.row, column=idx + 1)
            tgt.value = cell.value
            tgt.fill = cell.fill
        for row in ws.iter_rows(min_row=1, min_col=idx, max_col=idx):
            cell = row[0]
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
    ws.cell(row=1, column=new_col).value = "local_result"
    cols["local_result"] = new_col
    for name, idx in later_cols:
        cols[name] = idx + 1
    return new_col


def derive_xpu_classname(row_get) -> str | None:
    cn_xpu = row_get("classname_xpu")
    if cn_xpu:
        return str(cn_xpu).strip()
    cn_cuda = row_get("classname_cuda")
    if cn_cuda:
        return CUDA_SUFFIX_RE.sub("XPU", str(cn_cuda).strip())
    return None


def derive_test_name(row_get) -> str | None:
    n_xpu = row_get("name_xpu")
    if n_xpu:
        return str(n_xpu).strip()
    n_cuda = row_get("name_cuda")
    if n_cuda:
        return str(n_cuda).strip()
    return None


def resolve_test_file(testfile_cuda: str) -> tuple[Path, str] | None:
    """Return (absolute_file_path, location_label) or None.

    Lookup order (user-locked policy): pytorch/test/<full_subpath> first
    (preserves disambiguating subpath, so e.g. test/export/test_sparse.py
    is not confused with test/test_sparse.py). Only when the upstream
    pytorch file does not exist do we fall back to the torch-xpu-ops
    XPU mirror at torch-xpu-ops/test/xpu/[<subdir>/]<basename>_xpu.py.
    """
    if not testfile_cuda:
        return None
    rel = Path(testfile_cuda)
    # Excel `testfile_cuda` values include a leading `test/` segment
    # (e.g. `test/export/test_sparse.py`). PYTORCH_TEST_DIR already ends
    # in `/test`, so strip the redundant prefix before joining.
    if rel.parts and rel.parts[0] == "test":
        rel = Path(*rel.parts[1:]) if len(rel.parts) > 1 else Path(rel.name)
    primary = PYTORCH_TEST_DIR / rel
    if primary.is_file():
        return primary, "pytorch-test"
    basename = rel.stem
    suffix = rel.suffix or ".py"
    xpu_name = f"{basename}_xpu{suffix}"
    candidates_xpu = [
        XPU_OPS_TEST_DIR / rel.parent / xpu_name if rel.parent.parts else None,
        XPU_OPS_TEST_DIR / rel.parent.name / xpu_name if rel.parent.name else None,
        XPU_OPS_TEST_DIR / xpu_name,
    ]
    for c in candidates_xpu:
        if c and c.is_file():
            return c, "xpu-ops"
    return None


def safe_log_name(testfile: str, classname: str, name: str) -> str:
    s = f"{testfile}__{classname}__{name}"
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s)[:240] + ".log"


def is_distributed_test(testfile_cuda: str) -> bool:
    if not testfile_cuda:
        return False
    parts = Path(testfile_cuda).parts
    if parts and parts[0] == "test":
        parts = parts[1:]
    return bool(parts) and parts[0] == "distributed"


def atomic_save(wb, target: Path) -> None:
    # Atomic save: write to sibling temp file, fsync, os.replace.
    # Guards against truncation if the process is killed mid-save: the original
    # file stays untouched until os.replace() flips the dirent in one syscall.
    tmp = target.with_suffix(target.suffix + ".tmp_save")
    wb.save(tmp)
    fd = os.open(tmp, os.O_RDONLY)
    try:
        os.fsync(fd)
    finally:
        os.close(fd)
    os.replace(tmp, target)


# --- slim-load fast path -----------------------------------------------------
# Loading the full ~70MB workbook via openpyxl parses ~1GB of XML across 8
# sheets and takes ~3-4 minutes / ~8GB RSS, even though Step 0 only touches the
# two XPU-skipped sheets (~18MB of XML). The slim-load path builds a temporary
# 2-sheet workbook in-memory, lets openpyxl load THAT (~10s), and on save
# splices the modified sheet XML back into the original 8-sheet archive,
# preserving sheets 1/2/3/4/7/8 bit-for-bit.

_NS_SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"

_SlimMap = dict[str, str]


def build_slim_workbook(original: Path, keep_sheets: tuple[str, ...], slim_path: Path) -> _SlimMap:
    """Write a slim xlsx to slim_path containing only keep_sheets.

    Copies all archive parts EXCEPT the unkept sheet XMLs, and rewrites
    [Content_Types].xml + xl/workbook.xml + xl/_rels/workbook.xml.rels to
    reference only the kept sheets.

    Returns mapping {sheet_name: original_target} so save_slim_workbook can
    splice the edited XML back into the original archive.
    """
    keep_set = set(keep_sheets)
    name_to_orig_target: _SlimMap = {}
    rid_to_target: dict[str, str] = {}

    with zipfile.ZipFile(original) as src:
        rels_xml = src.read("xl/_rels/workbook.xml.rels")
        rels_root = ET.fromstring(rels_xml)
        for rel in rels_root.findall(f"{{{_NS_PKG_REL}}}Relationship"):
            tgt = rel.get("Target")
            if tgt and tgt.lstrip("/").startswith("xl/worksheets/"):
                rid_to_target[rel.get("Id")] = tgt.lstrip("/")

        wb_xml = src.read("xl/workbook.xml")
        wb_root = ET.fromstring(wb_xml)
        sheets_elem = wb_root.find(f"{{{_NS_SS}}}sheets")
        if sheets_elem is None:
            raise RuntimeError("workbook.xml has no <sheets> element")
        kept_sheet_elems: list[ET.Element] = []
        all_sheet_targets: set[str] = set()
        for sh in list(sheets_elem):
            name = sh.get("name")
            rid = sh.get(f"{{{_NS_R}}}id")
            tgt = rid_to_target.get(rid)
            if tgt:
                all_sheet_targets.add(tgt)
            if name in keep_set:
                kept_sheet_elems.append(sh)
                if tgt:
                    name_to_orig_target[name] = tgt

        if len(name_to_orig_target) != len(keep_set):
            missing = keep_set - set(name_to_orig_target)
            raise RuntimeError(f"slim-load: sheets not found in workbook: {missing}")

        for sh in list(sheets_elem):
            sheets_elem.remove(sh)
        for sh in kept_sheet_elems:
            sheets_elem.append(sh)
        slim_wb_xml = ET.tostring(wb_root, xml_declaration=True, encoding="UTF-8")

        kept_targets = set(name_to_orig_target.values())
        for rel in list(rels_root):
            tgt = rel.get("Target", "").lstrip("/")
            if tgt in all_sheet_targets and tgt not in kept_targets:
                rels_root.remove(rel)
        slim_rels_xml = ET.tostring(rels_root, xml_declaration=True, encoding="UTF-8")

        ct_xml = src.read("[Content_Types].xml")
        ct_root = ET.fromstring(ct_xml)
        for override in list(ct_root.findall(f"{{{_NS_CT}}}Override")):
            part = override.get("PartName", "").lstrip("/")
            if part in all_sheet_targets and part not in kept_targets:
                ct_root.remove(override)
        slim_ct_xml = ET.tostring(ct_root, xml_declaration=True, encoding="UTF-8")

        drop = (all_sheet_targets - kept_targets) | {
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "[Content_Types].xml",
        }
        with zipfile.ZipFile(slim_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for info in src.infolist():
                if info.filename in drop:
                    continue
                dst.writestr(info, src.read(info.filename))
            dst.writestr("xl/workbook.xml", slim_wb_xml)
            dst.writestr("xl/_rels/workbook.xml.rels", slim_rels_xml)
            dst.writestr("[Content_Types].xml", slim_ct_xml)

    return name_to_orig_target


def _deflate_raw(data: bytes, level: int = 6) -> bytes:
    # Raw DEFLATE (no zlib header/trailer) per RFC 1951, as required by the
    # zip file format (compression method 8).
    return zlib.compress(data, level)[2:-4]


def _write_raw_entry(
    dst: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    raw_bytes: bytes,
) -> None:
    # Write a ZipInfo + already-compressed bytes directly to dst.fp, bypassing
    # zipfile.writestr (which would recompress + clobber dst.start_dir between
    # writes, breaking subsequent raw appends).
    info.header_offset = dst.fp.tell()
    dst.fp.write(info.FileHeader(zip64=info.file_size >= 0xFFFFFFFF))
    dst.fp.write(raw_bytes)
    dst.filelist.append(info)
    dst.NameToInfo[info.filename] = info


def splice_slim_into_original(
    slim_path: Path,
    original: Path,
    name_to_orig_target: _SlimMap,
    output: Path,
) -> None:
    """Atomic write of `output` = original archive with sheets replaced by the
    edited versions from slim_path. Also replaces xl/styles.xml since openpyxl
    rewrites it whenever new fills (e.g. the ADD8E6 marker) are introduced;
    omitting it would erase the blue marks.

    Implementation pins both the unchanged passthrough parts and the replaced
    parts to raw-byte writes against dst.fp (manual local file header + raw
    DEFLATE for replacements). Mixing zipfile.writestr with raw writes is
    unsafe because writestr updates dst.start_dir mid-stream, causing the EOCD
    to be written in the middle of the file. End-to-end cost: ~0.1-0.5s vs
    ~16s for a naive decompress+recompress copy.
    """
    import struct

    with zipfile.ZipFile(slim_path) as slim:
        slim_wb_root = ET.fromstring(slim.read("xl/workbook.xml"))
        slim_rels_root = ET.fromstring(slim.read("xl/_rels/workbook.xml.rels"))
        slim_rid_to_target = {
            r.get("Id"): r.get("Target", "").lstrip("/")
            for r in slim_rels_root.findall(f"{{{_NS_PKG_REL}}}Relationship")
        }
        name_to_slim_target: dict[str, str] = {}
        slim_sheets_elem = slim_wb_root.find(f"{{{_NS_SS}}}sheets")
        for sh in slim_sheets_elem if slim_sheets_elem is not None else ():
            n = sh.get("name")
            rid = sh.get(f"{{{_NS_R}}}id")
            t = slim_rid_to_target.get(rid)
            if n in name_to_orig_target and t:
                name_to_slim_target[n] = t

        replacements: dict[str, bytes] = {}
        for name, orig_target in name_to_orig_target.items():
            slim_target = name_to_slim_target.get(name)
            if not slim_target:
                raise RuntimeError(f"splice: sheet {name!r} missing from slim archive")
            replacements[orig_target] = slim.read(slim_target)
        if "xl/styles.xml" in slim.namelist():
            replacements["xl/styles.xml"] = slim.read("xl/styles.xml")

        tmp = output.with_suffix(output.suffix + ".tmp_splice")
        with zipfile.ZipFile(original) as src:
            dst = zipfile.ZipFile(tmp, "w", allowZip64=True)
            try:
                seen: set[str] = set()
                for info in src.infolist():
                    if info.filename in replacements:
                        payload = replacements[info.filename]
                        compressed = _deflate_raw(payload)
                        new_info = zipfile.ZipInfo(
                            filename=info.filename, date_time=info.date_time
                        )
                        new_info.compress_type = zipfile.ZIP_DEFLATED
                        new_info.compress_size = len(compressed)
                        new_info.file_size = len(payload)
                        new_info.CRC = zlib.crc32(payload)
                        new_info.external_attr = info.external_attr
                        new_info.flag_bits = 0
                        _write_raw_entry(dst, new_info, compressed)
                        seen.add(info.filename)
                    else:
                        src.fp.seek(info.header_offset)
                        local_hdr = src.fp.read(30)
                        if local_hdr[:4] != b"PK\x03\x04":
                            raise RuntimeError(f"bad LFH for {info.filename}")
                        fname_len, extra_len = struct.unpack("<HH", local_hdr[26:30])
                        src.fp.read(fname_len + extra_len)
                        raw = src.fp.read(info.compress_size)
                        new_info = zipfile.ZipInfo(
                            filename=info.filename, date_time=info.date_time
                        )
                        new_info.compress_type = info.compress_type
                        new_info.compress_size = info.compress_size
                        new_info.file_size = info.file_size
                        new_info.CRC = info.CRC
                        new_info.external_attr = info.external_attr
                        # Clear bit 3 (streaming/data-descriptor flag): we
                        # have CRC + sizes upfront, so the local file header
                        # is self-describing.
                        new_info.flag_bits = info.flag_bits & ~0x08
                        _write_raw_entry(dst, new_info, raw)
                missing = set(replacements) - seen
                if missing:
                    raise RuntimeError(
                        f"splice: replacements not found in original archive: {missing}"
                    )
                # All raw writes done; tell zipfile.close() where the central
                # directory starts so it doesn't overwrite our payload.
                dst.start_dir = dst.fp.tell()
                dst._didModify = True
            finally:
                dst.close()

    fd = os.open(tmp, os.O_RDONLY)
    try:
        os.fsync(fd)
    finally:
        os.close(fd)
    os.replace(tmp, output)


def parse_status(returncode: int, output: str, timed_out: bool) -> str:
    if timed_out:
        return "TIMEOUT"
    if returncode in (-11, 139, -6, 134):
        return "SEGFAULT"
    has_passed = re.search(r"\bPASSED\b", output) is not None
    has_failed = re.search(r"\bFAILED\b", output) is not None
    has_error = re.search(r"\bERROR\b", output) is not None
    has_not_found = re.search(r"ERROR: not found:", output) is not None
    has_skipped = (
        re.search(r"\bSKIPPED\b", output) is not None
        or re.search(r"\bXFAIL\b", output) is not None
    )
    if returncode == 0 and has_passed and not (has_failed or has_error):
        return "PASS"
    if has_failed:
        return "FAIL"
    if has_not_found:
        return "NOT FOUND"
    if has_error:
        return "ERROR"
    if has_skipped:
        return "SKIP"
    return "DESELECTED"


def run_one(test_file: Path, classname: str, name: str, log_path: Path) -> tuple[str, Path]:
    node = f"{test_file}::{classname}::{name}"
    cmd = [sys.executable, "-m", "pytest", "-v", "--no-header", "-p", "no:cacheprovider", node]
    log_path.parent.mkdir(parents=True, exist_ok=True)
    timed_out = False
    try:
        result = subprocess.run(
            cmd,
            cwd=str(test_file.parent),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            timeout=PYTEST_TIMEOUT + 5,
            text=True,
        )
        out = result.stdout or ""
        rc = result.returncode
    except subprocess.TimeoutExpired as e:
        timed_out = True
        partial = e.output or b""
        if isinstance(partial, bytes):
            partial = partial.decode("utf-8", errors="replace")
        out = partial + f"\n[TIMEOUT after {PYTEST_TIMEOUT}s]\n"
        rc = -1
    header = (
        f"# cmd: {' '.join(cmd)}\n"
        f"# cwd: {test_file.parent}\n"
        f"# returncode: {rc}\n"
        f"# per_test_timeout: {PYTEST_TIMEOUT}s\n"
        f"# timed_out: {timed_out}\n"
        "----- pytest output -----\n"
    )
    log_path.write_text(header + out)
    return parse_status(rc, out, timed_out), log_path


def process_workbook(
    workbook_path: Path,
    sheets: tuple[str, ...],
    skip_distributed: bool = True,
    skip_status_xpu_set: bool = True,
    slim_load: bool = True,
    skip_patterns: tuple[str, ...] = (),
) -> None:
    log_dir_default = Path(f"/tmp/opencode/{workbook_path.stem}_local_verify")
    log_dir = Path(os.environ.get("LOG_DIR", str(log_dir_default)))
    log_dir.mkdir(parents=True, exist_ok=True)

    backup = workbook_path.with_suffix(workbook_path.suffix + ".bak_step0")
    if not backup.exists():
        print(f"[backup] {workbook_path} -> {backup}", flush=True)
        shutil.copy2(workbook_path, backup)

    # Slim-load fast path: load a synthetic 2-sheet workbook (~1MB / ~4s)
    # instead of the full 8-sheet original (~70MB / ~205s / ~8GB RSS).
    # Checkpoint saves write to the slim file then splice back into the
    # original. See build_slim_workbook + splice_slim_into_original.
    slim_path: Path | None = None
    slim_map: _SlimMap | None = None
    edit_target: Path
    if slim_load:
        slim_path = log_dir / f"{workbook_path.stem}.slim.xlsx"
        print(f"[slim ] building slim workbook (sheets={list(sheets)})", flush=True)
        t0 = time.time()
        slim_map = build_slim_workbook(workbook_path, sheets, slim_path)
        print(
            f"[slim ] built {slim_path.name} ({slim_path.stat().st_size/1e6:.1f} MB) "
            f"in {time.time()-t0:.1f}s",
            flush=True,
        )
        edit_target = slim_path
    else:
        edit_target = workbook_path

    print(f"[load ] {edit_target}", flush=True)
    t0 = time.time()
    wb = openpyxl.load_workbook(edit_target)
    print(f"[load ] done in {time.time()-t0:.1f}s", flush=True)

    def persist(label: str) -> None:
        # Save to edit_target (slim or original), then splice slim -> original
        # if we're on the fast path. The original on disk is always
        # checkpoint-consistent so a crashed run can be resumed directly.
        t1 = time.time()
        atomic_save(wb, edit_target)
        if slim_load and slim_path is not None and slim_map is not None:
            splice_slim_into_original(slim_path, backup, slim_map, workbook_path)
        print(f"[save ] {label} in {time.time()-t1:.1f}s", flush=True)

    pending_save = False
    processed_total = 0
    sheet_summaries: list[str] = []

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            print(f"[skip ] sheet {sheet_name!r} not in workbook", flush=True)
            continue
        ws = wb[sheet_name]
        cols = header_index(ws)
        for required in ("testfile_cuda", "classname_cuda", "name_cuda", "Reason", "DetailReason"):
            if required not in cols:
                raise RuntimeError(f"sheet {sheet_name!r} missing required column {required!r}")

        local_col = ensure_local_result_column(ws, cols)
        ws.cell(row=1, column=local_col).fill = BLUE

        reason_col = cols["Reason"]
        detail_col = cols["DetailReason"]

        per_sheet = 0
        per_sheet_skipped = 0
        statuses: dict[str, int] = {}

        for row_idx in range(2, ws.max_row + 1):
            if MAX_ROWS and per_sheet >= MAX_ROWS:
                break

            def row_get(col_name: str, _r=row_idx):
                idx = cols.get(col_name)
                return ws.cell(row=_r, column=idx).value if idx else None

            reason_val = row_get("Reason")
            local_val = ws.cell(row=row_idx, column=local_col).value
            if reason_val not in (None, ""):
                continue
            if local_val not in (None, ""):
                per_sheet_skipped += 1
                continue

            testfile_cuda = (row_get("testfile_cuda") or "").strip()
            if skip_distributed and is_distributed_test(testfile_cuda):
                per_sheet_skipped += 1
                continue
            if skip_status_xpu_set and row_get("status_xpu") not in (None, ""):
                per_sheet_skipped += 1
                continue
            classname = derive_xpu_classname(row_get)
            name = derive_test_name(row_get)
            if skip_patterns:
                hay = " ".join(
                    str(x) for x in (testfile_cuda, classname, name) if x
                ).lower()
                if any(p.lower() in hay for p in skip_patterns):
                    per_sheet_skipped += 1
                    continue
            if not (testfile_cuda and classname and name):
                cell = ws.cell(row=row_idx, column=local_col)
                cell.value = "ERROR;missing testfile/classname/name"
                cell.fill = BLUE
                pending_save = True
                per_sheet += 1
                statuses["ERROR"] = statuses.get("ERROR", 0) + 1
                continue

            resolved = resolve_test_file(testfile_cuda)
            if resolved is None:
                cell = ws.cell(row=row_idx, column=local_col)
                cell.value = (
                    f"ERROR;test file not found (looked in "
                    f"torch-xpu-ops/test/xpu/{Path(testfile_cuda).stem}_xpu.py "
                    f"and pytorch/test/{testfile_cuda})"
                )
                cell.fill = BLUE
                pending_save = True
                per_sheet += 1
                statuses["ERROR"] = statuses.get("ERROR", 0) + 1
                continue

            test_file, where = resolved
            log_path = log_dir / safe_log_name(testfile_cuda, classname, name)
            t_start = time.time()
            status, log_path = run_one(test_file, classname, name, log_path)
            elapsed = time.time() - t_start
            print(
                f"[{sheet_name[:24]:<24}] row {row_idx:>6} {where:>11} "
                f"{status:<8} {elapsed:5.1f}s  {classname}::{name}",
                flush=True,
            )

            cell_lr = ws.cell(row=row_idx, column=local_col)
            cell_lr.value = f"{status};{log_path}"
            cell_lr.fill = BLUE

            if status == "PASS":
                cell_r = ws.cell(row=row_idx, column=reason_col)
                cell_d = ws.cell(row=row_idx, column=detail_col)
                cell_r.value = "Local Passed"
                cell_d.value = (
                    f"{CONFIDENCE_PREFIX} Local pre-screen PASS on pytorch_opencode_env "
                    f"({where}); log: {log_path}"
                )
                cell_r.fill = BLUE
                cell_d.fill = BLUE

            pending_save = True
            per_sheet += 1
            processed_total += 1
            statuses[status] = statuses.get(status, 0) + 1

            if pending_save and (processed_total % CHECKPOINT_EVERY == 0):
                persist(f"checkpoint after {processed_total} rows")
                pending_save = False

        summary = (
            f"  {sheet_name}: processed {per_sheet}, "
            f"already-prescreened {per_sheet_skipped}, "
            + ", ".join(f"{k}={v}" for k, v in sorted(statuses.items()))
        )
        sheet_summaries.append(summary)
        print(summary, flush=True)

    if pending_save:
        persist("final save")
    else:
        print("[save ] nothing to save", flush=True)

    print("\n===== SUMMARY =====")
    for s in sheet_summaries:
        print(s)
    print(f"[done ] total processed rows: {processed_total}")
    print(f"[done ] logs:    {log_dir}")
    print(f"[done ] backup:  {backup}")


def main() -> int:
    global PYTEST_TIMEOUT
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("workbook", help="Path to .xlsx workbook")
    ap.add_argument(
        "--sheet",
        action="append",
        help="Sheet name (repeatable). Default: both XPU-skipped sheets.",
    )
    ap.add_argument(
        "--timeout",
        type=int,
        default=None,
        help=(
            "Per-test timeout in seconds. Overrides the PYTEST_PER_TEST_TIMEOUT "
            f"env var. Default: {PYTEST_TIMEOUT}."
        ),
    )
    ap.add_argument(
        "--include-distributed",
        action="store_true",
        help=(
            "Process distributed tests (testfile under test/distributed/...). "
            "Default: skip them. Distributed tests typically need multi-rank "
            "launchers (torchrun / mpiexec) and time out under single-process pytest."
        ),
    )
    ap.add_argument(
        "--include-status-xpu-set",
        action="store_true",
        help=(
            "Process rows whose status_xpu is already populated (skipped/failed/"
            "xfail). Default: only process rows with blank status_xpu, on the "
            "assumption that populated rows already reflect a known XPU outcome."
        ),
    )
    ap.add_argument(
        "--no-slim-load",
        action="store_true",
        help=(
            "Disable slim-load fast path; openpyxl loads the full workbook "
            "(~205s, ~8GB RSS on a 70MB / 8-sheet input). Default: slim-load "
            "ON (~4s init, ~3s/checkpoint via raw-zip splice)."
        ),
    )
    ap.add_argument(
        "--skip",
        action="append",
        default=[],
        metavar="PATTERN",
        help=(
            "Skip rows whose testfile_cuda, derived classname, or test name "
            "contains PATTERN (case-insensitive substring match). Repeatable. "
            "Example: --skip CPU skips rows like TestFooCPU; --skip CPU --skip "
            "Meta skips both. Skipped rows are NOT executed and NOT written."
        ),
    )
    args = ap.parse_args()
    wb_path = Path(args.workbook).resolve()
    if not wb_path.is_file():
        print(f"workbook not found: {wb_path}", file=sys.stderr)
        return 2
    if args.timeout is not None:
        if args.timeout <= 0:
            print(f"--timeout must be positive, got {args.timeout}", file=sys.stderr)
            return 2
        PYTEST_TIMEOUT = args.timeout
    sheets = tuple(args.sheet) if args.sheet else SHEETS
    process_workbook(
        wb_path,
        sheets,
        skip_distributed=not args.include_distributed,
        skip_status_xpu_set=not args.include_status_xpu_set,
        slim_load=not args.no_slim_load,
        skip_patterns=tuple(args.skip),
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
