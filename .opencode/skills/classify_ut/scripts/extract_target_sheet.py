#!/usr/bin/env python3
"""
Extract a single target sheet from a large status workbook into its own file.

The weekly UT status workbooks carry many sheets and thousands of rows, which
makes every downstream open/save slow. This step copies only the target sheet
into a new, small workbook; all later preparation and classification phases then
operate on that extracted file instead of the original.

The original workbook is never modified. The target sheet is preserved as-is
(values, fills, column widths) by loading the source once and deleting every
other sheet before saving the copy.

Usage:
    python3 extract_target_sheet.py <workbook.xlsx> --sheet "<Sheet Name>"
    python3 extract_target_sheet.py <workbook.xlsx> --sheet "<Sheet Name>" --out /path/out.xlsx

Prints the path of the extracted workbook on success.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl


def _slug(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z]+", "_", name).strip("_") or "sheet"


def extract_target_sheet(workbook: str, sheet: str, out: str | None = None) -> Path:
    src = Path(workbook).expanduser()
    if not src.is_file():
        raise SystemExit(f"Workbook not found: {src}")

    wb = openpyxl.load_workbook(src)
    if sheet not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {sheet!r} not found in {src.name}; available: {wb.sheetnames}"
        )

    for name in [n for n in wb.sheetnames if n != sheet]:
        del wb[name]
    wb.active = wb[sheet]

    dst = (
        Path(out).expanduser()
        if out
        else src.with_name(f"{src.stem}.{_slug(sheet)}.xlsx")
    )
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)
    return dst


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to the original (large) .xlsx.")
    parser.add_argument("--sheet", required=True, help="Target sheet name to extract.")
    parser.add_argument("--out", help="Output path (default: <stem>.<sheet>.xlsx).")
    args = parser.parse_args()

    dst = extract_target_sheet(args.workbook, args.sheet, args.out)
    ws = openpyxl.load_workbook(dst, read_only=True).active
    print(f"Extracted sheet '{args.sheet}' -> {dst}")
    print(f"  rows={ws.max_row} cols={ws.max_column}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
