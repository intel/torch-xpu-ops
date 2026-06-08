#!/usr/bin/env python3
"""
Filter rows of an extracted target-sheet workbook to a smaller subset workbook.

Used as the row-level filter step in the classify_ut workflow: after the target
sheet is extracted to a small workbook, this script applies a simple
column=value AND-filter to produce a subset workbook that the agent edits in
isolation. The subset workbook carries a `_source_row` column so the matching
cells can be written back to the source via `apply_filtered_changes.py`.

Filter syntax:
    --where "Column=value" ["Column2!=value2" "Column3~=substring" ...]

Operators:
    `=`   exact match (empty value matches blank cells)
    `!=`  inverse of `=`
    `~=`  cell value (as string) contains the given substring (non-empty)

Multiple `--where` tokens are AND-ed. Values that contain spaces must be
shell-quoted; the script does not re-tokenize.

Usage:
    # Typical "classify blank-Reason rows" use case
    python3 filter_target_rows.py extracted.xlsx --where "Reason=" \\
        --out blank_only.xlsx

    # TBE re-verification pass: only rows opted in
    python3 filter_target_rows.py extracted.xlsx \\
        --where "TBE_Reverify=True" \\
        --out tbd_reverify.xlsx

    # Subset by Reason and DetailReason (the example from the user)
    python3 filter_target_rows.py extracted.xlsx \\
        --where "Reason=To be enabled" "DetailReason~=Daisy" \\
        --out daisy_subset.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl


SOURCE_ROW_COLUMN = "_source_row"


def _parse_where(token: str) -> tuple[str, str, str]:
    # Order matters: `!=` and `~=` are two chars; `=` is a single char.
    # We check multi-char operators first, then fall back to `=`.
    for op_candidate in ("!=", "~="):
        if op_candidate in token:
            col, _, value = token.partition(op_candidate)
            return col.strip(), op_candidate, value
    if "=" in token:
        col, _, value = token.partition("=")
        return col.strip(), "=", value
    raise ValueError(f"--where token {token!r} must contain '=', '!=', or '~='")


def _row_matches(
    row_values: list,
    header: list[str],
    filters: list[tuple[str, str, str]],
) -> bool:
    row_dict = dict(zip(header, row_values))
    for col, op, value in filters:
        cell = row_dict.get(col)
        cell_blank = cell is None or (isinstance(cell, str) and cell.strip() == "")
        if op == "=":
            if value == "":
                if not cell_blank:
                    return False
            else:
                # String comparison avoids dtype surprises (int vs str in mixed columns).
                if cell is None or str(cell) != value:
                    return False
        elif op == "!=":
            if value == "":
                if cell_blank:
                    return False
            else:
                if cell is not None and str(cell) == value:
                    return False
        elif op == "~=":
            if value == "":
                if not cell_blank:
                    return False
            else:
                if cell is None:
                    return False
                if value not in str(cell):
                    return False
    return True


def filter_target_rows(
    workbook: str,
    where: list[str],
    out: str | None = None,
    add_source_row: bool = True,
) -> Path | None:
    """Filter the first sheet of `workbook` by the given --where tokens.

    Writes the matching rows to `out` (default: <stem>.filtered.xlsx), appending
    a `_source_row` column with the 1-based row number in the source.

    Returns the output path, or None if no rows matched.
    """
    src = Path(workbook).expanduser()
    if not src.is_file():
        raise SystemExit(f"Workbook not found: {src}")

    filters = [_parse_where(t) for t in where]

    src_wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    src_ws = src_wb.active
    if src_ws is None:
        raise SystemExit(f"No active sheet in {src}")

    header = [c.value for c in next(src_ws.iter_rows(min_row=1, max_row=1))]
    for col, _, _ in filters:
        if col not in header:
            raise SystemExit(
                f"Filter column {col!r} not found in {src.name}; "
                f"available columns: {[c for c in header if c]}"
            )

    matched_rows: list[tuple[int, list]] = []
    for r, row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)
        if _row_matches(row_values, header, filters):
            matched_rows.append((r, row_values))
    src_wb.close()

    # Column widths are not available in read_only mode; do a quick second pass.
    widths: dict[str, float] = {}
    try:
        width_wb = openpyxl.load_workbook(src, read_only=False)
        width_ws = width_wb.active
        if width_ws is not None and width_ws.column_dimensions:
            for col_letter, dim in width_ws.column_dimensions.items():
                if dim.width is not None:
                    widths[col_letter] = dim.width
        width_wb.close()
    except Exception:
        widths = {}

    dst = (
        Path(out).expanduser()
        if out
        else src.with_name(f"{src.stem}.filtered.xlsx")
    )

    if not matched_rows:
        return None

    out_header = list(header)
    if add_source_row and SOURCE_ROW_COLUMN not in out_header:
        out_header.append(SOURCE_ROW_COLUMN)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = src_ws.title
    for c, name in enumerate(out_header, start=1):
        out_ws.cell(row=1, column=c, value=name)
    for out_r, (src_r, row_values) in enumerate(matched_rows, start=2):
        for c, val in enumerate(row_values, start=1):
            out_ws.cell(row=out_r, column=c, value=val)
        if add_source_row and SOURCE_ROW_COLUMN not in header:
            out_ws.cell(row=out_r, column=len(out_header), value=src_r)

    for col_letter, width in widths.items():
        out_ws.column_dimensions[col_letter].width = width

    dst.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(dst)
    return dst


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("workbook", help="Path to the extracted-sheet .xlsx.")
    parser.add_argument(
        "--where",
        nargs="+",
        required=True,
        help=(
            "One or more Column=value or Column!=value tokens, AND-ed together. "
            "Empty value (e.g. 'Reason=') matches blank cells."
        ),
    )
    parser.add_argument("--out", help="Output path (default: <stem>.filtered.xlsx).")
    parser.add_argument(
        "--no-source-row",
        action="store_true",
        help=(
            "Do NOT add the _source_row column. Use this only when the output will "
            "not be fed to apply_filtered_changes.py."
        ),
    )
    args = parser.parse_args()

    dst = filter_target_rows(
        args.workbook,
        args.where,
        args.out,
        add_source_row=not args.no_source_row,
    )
    if dst is None:
        print("Filter matched 0 rows; no output file written.", file=sys.stderr)
        return 1
    out_ws = openpyxl.load_workbook(dst, read_only=True).active
    n_rows = (out_ws.max_row or 0) - 1
    n_cols = out_ws.max_column or 0
    print(f"Filtered -> {dst}")
    print(f"  matched rows: {n_rows}")
    print(f"  columns: {n_cols}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
