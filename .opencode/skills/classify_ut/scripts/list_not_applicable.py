#!/usr/bin/env python3
"""
List the not-target Operation/API entries used by the CUDA-Only Judgement Rule.

The authoritative source is the `Not applicable` sheet of the reference workbook
`torch_xpu_ops_issues.xlsx`. Its `Operation/API` column enumerates the operators
and APIs that are out of XPU scope (CUDA-only behavior, deprecated/removed
features, third-party-dependency gaps, etc.). Each row also carries the deciding
`Issue ID` and `Category`, which classification cites in `DetailReason`.

By default the workbook is fetched from its public GitHub location so callers do
not depend on any private checkout path. Pass `--xlsx PATH` (or set
`NOT_APPLICABLE_XLSX`) to read a local copy instead.

Usage:
    python3 list_not_applicable.py            # human-readable table
    python3 list_not_applicable.py --json     # JSON array for programmatic use

As a module:
    from list_not_applicable import load_not_applicable
    rows = load_not_applicable()              # list[dict]
"""

from __future__ import annotations

import argparse
import io
import json
import os
import sys
import urllib.request

import openpyxl

WORKBOOK_URL = os.environ.get(
    "NOT_APPLICABLE_XLSX_URL",
    "https://raw.githubusercontent.com/daisyden/ai_for_validation/main"
    "/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx",
)
SHEET_NAME = "Not applicable"
FIELDS = ("Issue ID", "Title", "Operation/API", "Category", "Labels", "State")


def _open_workbook_bytes(xlsx: str | None) -> bytes:
    if xlsx:
        with open(os.path.expanduser(xlsx), "rb") as fh:
            return fh.read()
    with urllib.request.urlopen(WORKBOOK_URL) as resp:
        return resp.read()


def load_not_applicable(xlsx: str | None = None) -> list[dict]:
    """Return `Not applicable` rows as dicts; `xlsx` local path else download."""
    wb = openpyxl.load_workbook(
        io.BytesIO(_open_workbook_bytes(xlsx)), read_only=True, data_only=True
    )
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {SHEET_NAME!r} not found; available: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {f: header.index(f) for f in FIELDS if f in header}
    out = []
    for raw in rows[1:]:
        if not any(raw):
            continue
        row = {f: (raw[idx[f]] if f in idx else None) for f in FIELDS}
        if not row.get("Operation/API"):
            continue
        out.append(row)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        default=os.environ.get("NOT_APPLICABLE_XLSX"),
        help="Local workbook path (default: download from NOT_APPLICABLE_XLSX_URL).",
    )
    parser.add_argument(
        "--json", action="store_true", help="Emit JSON instead of a table."
    )
    args = parser.parse_args()

    rows = load_not_applicable(args.xlsx)

    if args.json:
        json.dump(rows, sys.stdout, ensure_ascii=False, indent=2, default=str)
        sys.stdout.write("\n")
        return 0

    print(f"{len(rows)} not-applicable Operation/API entries:\n")
    for r in rows:
        print(f"- Issue #{r['Issue ID']} [{r['State']}] {r['Category']}")
        print(f"    Operation/API: {r['Operation/API']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
