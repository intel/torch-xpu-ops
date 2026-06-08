#!/usr/bin/env python3
"""
Apply Reason/DetailReason/etc. changes from a filtered subset workbook back to
the target workbook.

Used as the final step in the classify_ut workflow: the agent edits a
filtered subset (produced by `filter_target_rows.py`); this script then writes
the agent's decisions back to the target workbook, keyed by the `_source_row`
column that the filter script adds.

The script:
- Reads the filtered workbook.
- For each row, uses `_source_row` to find the matching 1-based row in the
  target workbook.
- Copies the values in the named `--write-columns` from the filtered row to
  the target row.
- Marks each updated cell blue (default: ADD8E6).
- Saves the target workbook in place (or to --out if specified).

Usage:
    # Round-trip: edits from filtered.xlsx go back into the extracted sheet.
    python3 apply_filtered_changes.py extracted.xlsx filtered.xlsx \\
        --write-columns Reason DetailReason

    # Write a new agent output without touching the source.
    python3 apply_filtered_changes.py source.xlsx filtered.xlsx \\
        --out agent.xlsx \\
        --write-columns Reason DetailReason Confidence
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


SOURCE_ROW_COLUMN = "_source_row"
DEFAULT_BLUE = "ADD8E6"


def _resolve_columns(
    header: list,
    columns: list[str],
    workbook_name: str,
    label: str,
) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in columns:
        if col not in header:
            raise SystemExit(
                f"{label} column {col!r} not found in {workbook_name}; "
                f"available columns: {[c for c in header if c]}"
            )
        out[col] = header.index(col) + 1
    return out


def apply_filtered_changes(
    target: str,
    filtered: str,
    write_columns: list[str],
    from_column: str = SOURCE_ROW_COLUMN,
    blue: str = DEFAULT_BLUE,
    out: str | None = None,
) -> dict[str, int]:
    """Apply changes from the filtered workbook to the target workbook.

    Returns a summary dict: {column: cells_updated, "_total": N}.
    """
    target_path = Path(target).expanduser()
    filtered_path = Path(filtered).expanduser()
    if not target_path.is_file():
        raise SystemExit(f"Target workbook not found: {target_path}")
    if not filtered_path.is_file():
        raise SystemExit(f"Filtered workbook not found: {filtered_path}")

    target_wb = openpyxl.load_workbook(target_path)
    target_ws = target_wb.active
    if target_ws is None:
        raise SystemExit(f"No active sheet in {target_path}")

    target_header = [target_ws.cell(row=1, column=c).value for c in range(1, target_ws.max_column + 1)]
    target_col_idx = _resolve_columns(target_header, write_columns, target_path.name, "Write")

    filtered_wb = openpyxl.load_workbook(filtered_path)
    filtered_ws = filtered_wb.active
    if filtered_ws is None:
        raise SystemExit(f"No active sheet in {filtered_path}")

    filtered_header = [filtered_ws.cell(row=1, column=c).value for c in range(1, filtered_ws.max_column + 1)]
    if from_column not in filtered_header:
        raise SystemExit(
            f"From-column {from_column!r} not found in filtered workbook; "
            f"available columns: {[c for c in filtered_header if c]}"
        )
    from_col_idx = filtered_header.index(from_column) + 1
    filtered_col_idx = _resolve_columns(filtered_header, write_columns, filtered_path.name, "Write")

    blue_fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
    summary = {col: 0 for col in write_columns}
    summary["_total"] = 0

    for r in range(2, filtered_ws.max_row + 1):
        source_row = filtered_ws.cell(row=r, column=from_col_idx).value
        if not isinstance(source_row, int):
            raise SystemExit(
                f"Row {r} in filtered workbook has invalid {from_column}: {source_row!r}"
            )
        if source_row < 2 or source_row > target_ws.max_row:
            raise SystemExit(
                f"Row {r} {from_column}={source_row} is outside the target "
                f"row range [2, {target_ws.max_row}]"
            )
        for col in write_columns:
            new_val = filtered_ws.cell(row=r, column=filtered_col_idx[col]).value
            target_cell = target_ws.cell(row=source_row, column=target_col_idx[col])
            if target_cell.value != new_val:
                target_cell.value = new_val
                target_cell.fill = blue_fill
                summary[col] += 1
                summary["_total"] += 1

    out_path = Path(out).expanduser() if out else target_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(out_path)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("target", help="Target workbook to write into (or copy from if --out is set).")
    parser.add_argument("filtered", help="Filtered workbook with the agent's edits.")
    parser.add_argument(
        "--write-columns",
        nargs="+",
        required=True,
        help="One or more column names whose values get copied from filtered -> target.",
    )
    parser.add_argument(
        "--from-column",
        default=SOURCE_ROW_COLUMN,
        help=f"Column carrying the source row number (default: {SOURCE_ROW_COLUMN}).",
    )
    parser.add_argument(
        "--blue",
        default=DEFAULT_BLUE,
        help=f"Fill color for updated cells (default: {DEFAULT_BLUE}).",
    )
    parser.add_argument(
        "--out",
        help="Output path (default: overwrite target in place).",
    )
    args = parser.parse_args()

    summary = apply_filtered_changes(
        args.target,
        args.filtered,
        args.write_columns,
        args.from_column,
        args.blue,
        args.out,
    )
    out_path = args.out or args.target
    print(f"Applied filtered changes -> {out_path}")
    for col, n in summary.items():
        if col == "_total":
            continue
        if n > 0:
            print(f"  {col}: {n} cells updated")
    print(f"  total: {summary['_total']} cells updated")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
