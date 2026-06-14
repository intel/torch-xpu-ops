#!/usr/bin/env python3
"""
write_results.py — Write classification results to the "agent" sheet of an Excel file.

Phase 3 of classify_ut workflow. Two modes:

BUILD (default):
  Reads the original input sheet and a results.json, then writes a fresh "agent"
  sheet (original columns + Analyzed, Reason, DetailReason, ReuseSource, Confidence).
  Use for the FIRST run, when the output file does not yet exist.

MERGE (--merge):
  Updates an EXISTING agent_results.xlsx in place. Only the rows present in
  results.json are touched (matched by CUDA identity: testfile_cuda/classname_cuda/
  name_cuda). Every other row -- including rows analyzed by previous runs -- is left
  byte-for-byte untouched. Use this for incremental runs that add to an accumulator.

Safety:
  - BUILD refuses to overwrite an existing output file whose "agent" sheet already
    has Analyzed=TRUE rows that are NOT in results.json (those rows would be lost).
    Pass --merge to update in place, or --force to overwrite anyway.
  - Any in-place write (MERGE, or BUILD --force over an existing file) first writes
    a <file>.bak backup.

Confidence is auto-computed: High if DetailReason contains exact evidence
(commit hash, issue/PR URL, PR reference), otherwise Medium.

Usage:
    # First run (build a fresh standalone output file):
    python3 write_results.py <excel_path> <results.json> [sheet_name] \
        --output_sheet=agent --output-excel=agent_results.xlsx

    # Incremental run (safely merge into the existing accumulator):
    python3 write_results.py --merge <results.json> \
        --output_sheet=agent --output-excel=agent_results.xlsx
"""

import json
import os
import re
import shutil
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


RESULT_COLUMNS = ["Analyzed", "Reason", "DetailReason", "ReuseSource", "Confidence"]
IDENTITY_COLUMNS = ["testfile_cuda", "classname_cuda", "name_cuda"]

EVIDENCE_PATTERNS = [
    re.compile(r'github\.com/[\w.-]+/[\w.-]+/issues/\d+'),
    re.compile(r'github\.com/[\w.-]+/[\w.-]+/pull/\d+'),
    re.compile(r'\b[0-9a-f]{7,40}\b'),
    re.compile(r'#\d+'),
]


def compute_confidence(detail_reason: str) -> str:
    if not detail_reason:
        return "Medium"
    for pat in EVIDENCE_PATTERNS:
        if pat.search(detail_reason):
            return "High"
    return "Medium"


def identity_key(record, fields):
    vals = []
    have_any = False
    for f in fields:
        v = record.get(f)
        if v is not None and str(v) != "":
            have_any = True
        vals.append("" if v is None else str(v))
    return tuple(vals) if have_any else None


def backup_file(path):
    bak = path + ".bak"
    shutil.copy2(path, bak)
    print(f"Backup written: {bak}")


def parse_args(argv):
    opts = {"output_sheet": "agent", "output_excel": None, "force": False, "merge": False}
    positionals = []
    for arg in argv:
        if arg == "--merge":
            opts["merge"] = True
        elif arg == "--force":
            opts["force"] = True
        elif arg.startswith("--output_sheet="):
            opts["output_sheet"] = arg.split("=", 1)[1]
        elif arg.startswith("--output-excel="):
            opts["output_excel"] = arg.split("=", 1)[1]
        elif arg.startswith("--"):
            print(f"ERROR: unknown flag {arg}", file=sys.stderr)
            sys.exit(1)
        else:
            positionals.append(arg)
    return ("merge" if opts["merge"] else "build"), positionals, opts


def load_results(results_path):
    with open(results_path) as f:
        results = json.load(f)
    if isinstance(results, dict) and "results" in results:
        results = results["results"]
    return results


def header_index(headers, name):
    return headers.index(name) if name in headers else -1


def do_merge(results_path, opts):
    output_excel = opts["output_excel"]
    output_sheet = opts["output_sheet"]
    if not output_excel:
        print("ERROR: --merge requires --output-excel=<existing agent_results.xlsx>", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(output_excel):
        print(f"ERROR: --merge target '{output_excel}' does not exist. Run a build first "
              f"(without --merge) to create it.", file=sys.stderr)
        sys.exit(1)

    results = load_results(results_path)

    wb = load_workbook(output_excel)
    if output_sheet not in wb.sheetnames:
        print(f"ERROR: sheet '{output_sheet}' not found in {output_excel}. Sheets: {wb.sheetnames}",
              file=sys.stderr)
        sys.exit(1)
    ws = wb[output_sheet]
    headers = [c.value for c in ws[1]]

    missing = [c for c in RESULT_COLUMNS if c not in headers]
    if missing:
        print(f"ERROR: output sheet '{output_sheet}' is missing result columns {missing}. "
              f"It does not look like a built agent sheet; run a build first.", file=sys.stderr)
        sys.exit(1)

    col = {c: headers.index(c) + 1 for c in RESULT_COLUMNS}

    key_fields = [f for f in IDENTITY_COLUMNS if f in headers]
    if not key_fields:
        print("ERROR: output sheet has none of the CUDA identity columns "
              f"{IDENTITY_COLUMNS}; cannot match rows.", file=sys.stderr)
        sys.exit(1)

    name_idx = header_index(headers, "name_cuda")
    field_pos = {f: headers.index(f) for f in key_fields}
    full_index = {}
    name_index = {}
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = {f: row[field_pos[f]] for f in key_fields}
        k = identity_key(rec, key_fields)
        if k is not None:
            full_index.setdefault(k, []).append(r_i)
        if name_idx >= 0 and row[name_idx] is not None:
            name_index.setdefault(str(row[name_idx]), []).append(r_i)

    analyzed_pos1 = col["Analyzed"]
    updated, skipped_analyzed, unmatched, ambiguous = 0, [], [], []
    for r in results:
        k = identity_key(r, key_fields)
        rows = full_index.get(k) if k is not None else None
        label = "/".join(str(r.get(f, "")) for f in key_fields)
        if not rows:
            nm = r.get("name_cuda") or r.get("name_xpu")
            rows = name_index.get(str(nm)) if nm is not None else None
            label = str(nm)
        if not rows:
            unmatched.append(label)
            continue
        if len(rows) > 1:
            ambiguous.append((label, rows))
            continue
        rownum = rows[0]
        cur = str(ws.cell(row=rownum, column=analyzed_pos1).value or "").strip().lower()
        if cur == "true" and not opts["force"]:
            skipped_analyzed.append(label)
            continue
        detail = r.get("DetailReason", "")
        ws.cell(row=rownum, column=col["Analyzed"], value="TRUE")
        ws.cell(row=rownum, column=col["Reason"], value=r.get("Reason", ""))
        ws.cell(row=rownum, column=col["DetailReason"], value=detail)
        ws.cell(row=rownum, column=col["ReuseSource"], value=r.get("ReuseSource", "") or "")
        ws.cell(row=rownum, column=col["Confidence"], value=compute_confidence(detail))
        updated += 1

    backup_file(output_excel)
    wb.save(output_excel)

    print(f"MERGE complete: {output_excel} (sheet '{output_sheet}')")
    print(f"  updated:          {updated}")
    print(f"  skipped (already Analyzed=TRUE, use --force to overwrite): {len(skipped_analyzed)}")
    print(f"  unmatched (not found in sheet): {len(unmatched)}")
    if unmatched:
        for u in unmatched[:20]:
            print(f"    - {u}")
    if ambiguous:
        print(f"  AMBIGUOUS (identity matched multiple rows, NOT written): {len(ambiguous)}")
        for label, rows in ambiguous[:20]:
            print(f"    - {label} -> rows {rows}")
        sys.exit(3)


def guard_existing_output(output_excel, output_sheet, results, force):
    if not output_excel or not os.path.exists(output_excel) or force:
        return
    existing = load_workbook(output_excel, read_only=True)
    try:
        if output_sheet not in existing.sheetnames:
            return
        ews = existing[output_sheet]
        eheaders = [c.value for c in next(ews.iter_rows(min_row=1, max_row=1))]
        ai = header_index(eheaders, "Analyzed")
        if ai < 0:
            return
        key_fields = [f for f in IDENTITY_COLUMNS if f in eheaders]
        field_pos = {f: eheaders.index(f) for f in key_fields}
        name_i = header_index(eheaders, "name_cuda")
        result_keys = set()
        result_names = set()
        for r in results:
            k = identity_key(r, key_fields) if key_fields else None
            if k is not None:
                result_keys.add(k)
            nm = r.get("name_cuda") or r.get("name_xpu")
            if nm is not None:
                result_names.add(str(nm))
        lost = 0
        for row in ews.iter_rows(min_row=2, values_only=True):
            if str(row[ai] or "").strip().lower() != "true":
                continue
            rec = {f: row[field_pos[f]] for f in key_fields}
            k = identity_key(rec, key_fields) if key_fields else None
            nm = str(row[name_i]) if name_i >= 0 and row[name_i] is not None else None
            if (k is not None and k in result_keys) or (nm is not None and nm in result_names):
                continue
            lost += 1
        if lost:
            print(
                f"ERROR: '{output_excel}' (sheet '{output_sheet}') already contains {lost} "
                f"already-analyzed row(s) that are NOT in {os.path.basename('results.json')}.\n"
                f"       A fresh build would discard them.\n"
                f"       -> Use --merge to update those results in place (preserves prior rows), or\n"
                f"       -> Use --force to overwrite the file anyway (a .bak backup is kept).",
                file=sys.stderr,
            )
            sys.exit(2)
    finally:
        existing.close()


def do_build(excel_path, results_path, sheet_name, opts):
    output_sheet = opts["output_sheet"]
    output_excel = opts["output_excel"]
    results = load_results(results_path)

    guard_existing_output(output_excel, output_sheet, results, opts["force"])

    result_map = {}
    for r in results:
        key = r.get("name_cuda", "") or r.get("name_xpu", "")
        if key:
            result_map[key] = r

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data_rows.append(row)

    output_headers = list(headers) + RESULT_COLUMNS

    if output_excel:
        from openpyxl import Workbook
        if os.path.exists(output_excel) and opts["force"]:
            backup_file(output_excel)
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = output_sheet
    else:
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        out_ws = wb.create_sheet(title=output_sheet)

    for col_idx, h in enumerate(output_headers, start=1):
        out_ws.cell(row=1, column=col_idx, value=h)

    name_col = header_index(headers, "name_cuda")
    analyzed_col_orig = header_index(headers, "Analyzed")
    reason_col_orig = header_index(headers, "Reason")
    detail_col_orig = header_index(headers, "DetailReason")

    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            out_ws.cell(row=row_idx, column=col_idx, value=val)

        name = str(row[name_col]) if 0 <= name_col < len(row) and row[name_col] is not None else ""
        result = result_map.get(name, {})

        ac = len(headers) + 1
        rc = len(headers) + 2
        dc = len(headers) + 3
        rsc = len(headers) + 4
        cc = len(headers) + 5

        orig_analyzed = ""
        if 0 <= analyzed_col_orig < len(row) and row[analyzed_col_orig] is not None:
            orig_analyzed = str(row[analyzed_col_orig]).lower()

        if orig_analyzed == "true":
            out_ws.cell(row=row_idx, column=ac, value="TRUE")
            orig_reason = str(row[reason_col_orig]) if 0 <= reason_col_orig < len(row) and row[reason_col_orig] is not None else ""
            orig_detail = str(row[detail_col_orig]) if 0 <= detail_col_orig < len(row) and row[detail_col_orig] is not None else ""
            out_ws.cell(row=row_idx, column=rc, value=orig_reason)
            out_ws.cell(row=row_idx, column=dc, value=orig_detail)
            out_ws.cell(row=row_idx, column=rsc, value="")
            out_ws.cell(row=row_idx, column=cc, value=compute_confidence(orig_detail))
        elif result:
            out_ws.cell(row=row_idx, column=ac, value="TRUE")
            out_ws.cell(row=row_idx, column=rc, value=result.get("Reason", ""))
            detail = result.get("DetailReason", "")
            out_ws.cell(row=row_idx, column=dc, value=detail)
            out_ws.cell(row=row_idx, column=rsc, value=result.get("ReuseSource", ""))
            out_ws.cell(row=row_idx, column=cc, value=compute_confidence(detail))
        else:
            out_ws.cell(row=row_idx, column=ac, value="FALSE")
            out_ws.cell(row=row_idx, column=rc, value="")
            out_ws.cell(row=row_idx, column=dc, value="")
            out_ws.cell(row=row_idx, column=rsc, value="")
            out_ws.cell(row=row_idx, column=cc, value="")

    if output_excel:
        out_wb.save(output_excel)
        print(f"Written to '{output_excel}' (sheet '{output_sheet}')")
    else:
        wb.save(excel_path)
        print(f"Written to sheet '{output_sheet}' in {excel_path}")


def main():
    mode, positionals, opts = parse_args(sys.argv[1:])

    if mode == "merge":
        if len(positionals) != 1:
            print("Usage: python3 write_results.py --merge <results.json> "
                  "--output_sheet=agent --output-excel=agent_results.xlsx", file=sys.stderr)
            sys.exit(1)
        do_merge(positionals[0], opts)
        return

    if len(positionals) < 2:
        print("Usage: python3 write_results.py <excel_path> <results.json> [sheet_name] "
              "[--output_sheet=agent] [--output-excel=output.xlsx] [--force]\n"
              "   or: python3 write_results.py --merge <results.json> "
              "--output-excel=agent_results.xlsx [--output_sheet=agent] [--force]", file=sys.stderr)
        sys.exit(1)

    excel_path = positionals[0]
    results_path = positionals[1]
    sheet_name = positionals[2] if len(positionals) > 2 else None
    do_build(excel_path, results_path, sheet_name, opts)


if __name__ == "__main__":
    main()
