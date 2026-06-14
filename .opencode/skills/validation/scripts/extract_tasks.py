#!/usr/bin/env python3
"""
extract_tasks.py — Extract rows needing classification from an Excel sheet.

Phase 1 of classify_ut workflow:
1. Reads an Excel sheet with columns: testfile_cuda, classname_cuda, name_cuda,
   message_xpu, status_xpu. Optionally also testfile_xpu, classname_xpu, name_xpu.
2. Computes XPU columns from CUDA columns when XPU columns are absent or blank:
   - testfile_xpu = testfile_cuda
   - classname_xpu = classname_cuda with CUDA suffix → XPU (e.g. TestFooCUDA → TestFooXPU)
   - name_xpu = name_cuda with _cuda → _xpu (e.g. test_foo_cuda_float32 → test_foo_xpu_float32)
3. Deduplicates rows that share the same classname_xpu + similar error message as an already-analyzed row
4. Outputs tasks.json (rows needing classification) and already_resolved.json (deduplicated rows)

Filter options (applied after dedup — only matching rows appear in the output):
  --filter-reason PATTERN        Include only rows where Reason matches (substring, case-insensitive)
  --filter-detailreason PATTERN  Include only rows where DetailReason contains this substring (case-insensitive)
  --filter-status-xpu VALUE      Include only rows where status_xpu equals this value

Usage:
    python3 extract_tasks.py <excel_path> [sheet_name] > tasks.json
    python3 extract_tasks.py <excel_path> [sheet_name] --filter-reason "To be enabled" --filter-detailreason "Daisy" --filter-status-xpu "" > tasks.json
"""

import json
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract_operators(message: str) -> set:
    """Extract operator/API references from an error message."""
    if not message:
        return set()
    ops = set()
    for m in re.finditer(r'aten::(\w+)', message):
        ops.add(f'aten::{m.group(1)}')
    for m in re.finditer(r'torch\.(\w+(?:\.\w+)*)', message):
        ops.add(f'torch.{m.group(1)}')
    for kw in ['CUDA error', 'cuBLAS', 'cuDNN', 'NCCL', 'XPU', 'SYCL']:
        if kw.lower() in message.lower():
            ops.add(kw)
    return ops


def normalize_message(msg: str) -> str:
    """Normalize a message for similarity comparison."""
    if not msg:
        return ""
    msg = msg.lower().strip()
    msg = re.sub(r'0x[0-9a-f]+', '', msg)
    msg = re.sub(r'/\S+', '', msg)
    msg = re.sub(r'\b\d{10,}\b', '', msg)
    msg = re.sub(r'\s+', ' ', msg).strip()
    return msg


def levenshtein_similarity(a: str, b: str) -> float:
    """Compute Levenshtein similarity (0.0 to 1.0), capped at 200 chars."""
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    a, b = a[:200], b[:200]
    n, m = len(a), len(b)
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    for i in range(n + 1):
        dp[i][0] = i
    for j in range(m + 1):
        dp[0][j] = j
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1, dp[i - 1][j - 1] + cost)
    max_len = max(n, m)
    return 1.0 - (dp[n][m] / max_len) if max_len > 0 else 1.0


def messages_similar(msg_a: str, msg_b: str, threshold: float = 0.7) -> bool:
    """Check if two error messages are similar (share operators OR high string similarity)."""
    ops_a = extract_operators(msg_a)
    ops_b = extract_operators(msg_b)
    if ops_a and ops_b and ops_a & ops_b:
        return True
    sim = levenshtein_similarity(normalize_message(msg_a), normalize_message(msg_b))
    return sim >= threshold


def infer_xpu_fields(row: dict) -> dict:
    out = dict(row)
    testfile_cuda = row.get("testfile_cuda", "")
    classname_cuda = row.get("classname_cuda", "")
    name_cuda = row.get("name_cuda", "")

    raw = row.get("testfile_xpu", "")
    out["testfile_xpu"] = raw if raw else testfile_cuda

    raw = row.get("classname_xpu", "")
    if raw:
        out["classname_xpu"] = raw
    elif classname_cuda.endswith("CUDA"):
        out["classname_xpu"] = classname_cuda[:-4] + "XPU"
    else:
        out["classname_xpu"] = classname_cuda

    raw = row.get("name_xpu", "")
    out["name_xpu"] = raw if raw else name_cuda.replace("_cuda", "_xpu")

    return out


def parse_filters(args):
    filters = {}
    i = 1
    while i < len(args):
        if args[i] == "--filter-reason" and i + 1 < len(args):
            filters["reason"] = args[i + 1]
            print("WARNING: --filter-reason filters by the OUTPUT Reason column. "
                  "This is for post-hoc subset extraction only. "
                  "Do NOT use it as a classification shortcut — "
                  "the cascade (Gates 1-4 via subagents) must still be run.",
                  file=sys.stderr)
            i += 2
        elif args[i] == "--filter-detailreason" and i + 1 < len(args):
            filters["detailreason"] = args[i + 1]
            print("WARNING: --filter-detailreason filters by the OUTPUT DetailReason column. "
                  "This is for post-hoc subset extraction only. "
                  "Do NOT use it as a classification shortcut — "
                  "the cascade (Gates 1-4 via subagents) must still be run.",
                  file=sys.stderr)
            i += 2
        elif args[i] == "--filter-status-xpu" and i + 1 < len(args):
            filters["status_xpu"] = args[i + 1]
            i += 2
        else:
            i += 1
    return filters


def row_matches_filters(row: dict, filters: dict) -> bool:
    if "reason" in filters:
        val = row.get("Reason", "")
        pattern = filters["reason"].lower()
        if pattern not in val.lower():
            return False
    if "detailreason" in filters:
        val = row.get("DetailReason", "")
        pattern = filters["detailreason"].lower()
        if pattern not in val.lower():
            return False
    if "status_xpu" in filters:
        val = row.get("status_xpu", "")
        expected = filters["status_xpu"]
        if val != expected:
            return False
    return True


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_tasks.py <excel_path> [sheet_name] [--filter-reason PATTERN] [--filter-detailreason PATTERN] [--filter-status-xpu VALUE]", file=sys.stderr)
        sys.exit(1)

    filters = parse_filters(sys.argv)

    # Consume positional args (excel_path, sheet_name) by skipping filter flags
    positional = []
    i = 1
    while i < len(sys.argv):
        if sys.argv[i].startswith("--filter-"):
            i += 2  # skip flag and its value
        else:
            positional.append(sys.argv[i])
            i += 1

    if not positional:
        print("ERROR: Missing excel_path", file=sys.stderr)
        sys.exit(1)

    excel_path = positional[0]
    sheet_name = positional[1] if len(positional) > 1 else None

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    required = ["testfile_cuda", "classname_cuda", "name_cuda"]
    missing = [c for c in required if c not in col_map]
    if missing:
        print(f"ERROR: Missing columns: {missing}", file=sys.stderr)
        sys.exit(1)

    has_status = "status_xpu" in col_map
    has_reason = "Reason" in col_map
    has_detail = "DetailReason" in col_map
    has_analyzed = "Analyzed" in col_map

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        entry = {"_row": row_idx}
        for col_name, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            entry[col_name] = str(val) if val is not None else ""
        entry = infer_xpu_fields(entry)
        rows.append(entry)

    analyzed_rows = [r for r in rows if has_analyzed and r.get("Analyzed", "").lower() == "true"]

    def make_task_entry(r):
        return {
            "testfile_cuda": r.get("testfile_cuda", ""),
            "classname_cuda": r.get("classname_cuda", ""),
            "name_cuda": r.get("name_cuda", ""),
            "testfile_xpu": r.get("testfile_xpu", ""),
            "classname_xpu": r.get("classname_xpu", ""),
            "name_xpu": r.get("name_xpu", ""),
            "message_xpu": r.get("message_xpu", ""),
        }

    tasks = []
    already_resolved = []

    for r in rows:
        already_analyzed = has_analyzed and r.get("Analyzed", "").lower() == "true"
        if already_analyzed:
            e = make_task_entry(r)
            e["Analyzed"] = True
            e["Reason"] = r.get("Reason", "")
            e["DetailReason"] = r.get("DetailReason", "")
            e["ReuseSource"] = ""
            already_resolved.append(e)
            continue

        cls_xpu = r.get("classname_xpu", "")
        msg = r.get("message_xpu", "")
        reused = False
        for ar in analyzed_rows:
            if ar.get("classname_xpu", "") == cls_xpu and messages_similar(ar.get("message_xpu", ""), msg):
                e = make_task_entry(r)
                e["Analyzed"] = True
                e["Reason"] = f"[Reused row#{ar.get('_row', '??')}] {ar.get('Reason', '')}"
                e["DetailReason"] = f"[Reused row#{ar.get('_row', '??')}] {ar.get('DetailReason', '')}"
                e["ReuseSource"] = ar.get("name_xpu", "")
                already_resolved.append(e)
                reused = True
                break

        if reused:
            continue

        if filters and not row_matches_filters(r, filters):
            continue

        e = make_task_entry(r)
        e["status_xpu"] = r.get("status_xpu", "")
        e["Reason"] = r.get("Reason", "")
        e["DetailReason"] = r.get("DetailReason", "")
        tasks.append(e)

    output = {
        "tasks": tasks,
        "already_resolved": already_resolved,
        "summary": {
            "total_rows": len(rows),
            "already_analyzed": sum(1 for r in rows if has_analyzed and r.get("Analyzed", "").lower() == "true"),
            "deduplicated": len(already_resolved) - sum(1 for r in rows if has_analyzed and r.get("Analyzed", "").lower() == "true"),
            "needs_classification": len(tasks),
        },
    }

    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
